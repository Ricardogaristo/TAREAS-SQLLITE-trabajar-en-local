"""
consolidar_alumnos.py
=====================
Consolida las dos fuentes de datos de tu sistema en un único DataFrame:

  • Fuente académica  → curso, nombre, examenes, progreso  (tabla `alumnos`)
  • Fuente de contacto → nombre, telefono               (misma tabla, campo separado)

El cruce se hace por nombre normalizado (sin tildes, minúsculas, sin espacios
extra), igual que hace tu función `norm()` en formacion.py.

CÓMO USAR:
    # Opción A — desde la BD directamente (lo más común)
    from consolidar_alumnos import consolidar_desde_db
    df = consolidar_desde_db(tutor_id=session["user_id"])

    # Opción B — desde el mismo Excel de importación
    from consolidar_alumnos import consolidar_desde_excel
    df = consolidar_desde_excel("alumnos.xlsx")

    # Opción C — desde listas en memoria (tests, scripts puntuales)
    from consolidar_alumnos import consolidar
    df = consolidar(lista_academica, lista_contacto)

COLUMNAS DEL DATAFRAME RESULTANTE:
    nombre_curso | nombre_alumno | telefono | examenes | progreso_general
"""

from __future__ import annotations

import io
import sqlite3
import unicodedata
from typing import Any

import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZACIÓN  (idéntica a tu `norm()` de formacion.py)
# ─────────────────────────────────────────────────────────────────────────────

def _norm(texto: Any) -> str:
    """Elimina tildes, pasa a minúsculas y colapsa espacios. Igual que norm() en formacion.py."""
    if texto is None:
        return ""
    s = str(texto).strip()
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(sin_tildes.lower().split())


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS INTERNOS  (misma lógica que _safe_float / _safe_int de formacion.py)
# ─────────────────────────────────────────────────────────────────────────────

def _safe_float(val: Any) -> float:
    try:
        return float(str(val).replace("%", "").replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_int(val: Any) -> int:
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0


def _limpiar_telefono(val: Any) -> str | None:
    """Mismo tratamiento que en tu importación: elimina .0, espacios y valores vacíos."""
    if val is None:
        return None
    s = str(val).strip().replace(".0", "").replace(" ", "")
    return s if s and s.lower() not in ("none", "nan", "") else None


# ─────────────────────────────────────────────────────────────────────────────
# DETECCIÓN AUTOMÁTICA DE COLUMNAS  (mismos alias que get_col() en formacion.py)
# ─────────────────────────────────────────────────────────────────────────────

_ALIAS_ACADEMICOS = {
    "nombre_curso":     ["del curso", "nombre del curso", "curso", "materia", "asignatura"],
    "nombre_alumno":    ["nombre", "alumno", "estudiante"],
    "progreso_general": ["progreso", "avance", "progress"],
    "examenes":         ["examen", "exam", "evaluac", "prueba"],
}

_ALIAS_CONTACTO = {
    "nombre_alumno": ["nombre", "alumno", "estudiante"],
    "telefono":      ["telefono", "celular", "phone", "whatsapp", "movil", "tel"],
}


def _detectar_columnas(headers: list[str], alias_map: dict, excluir: dict | None = None) -> dict[str, int | None]:
    """
    Devuelve {campo_destino: indice_columna}.
    Coincidencia parcial, mismo criterio que get_col() de formacion.py.
    """
    headers_norm = [_norm(h) for h in headers]
    resultado: dict[str, int | None] = {}
    excluir = excluir or {}

    for campo, posibles in alias_map.items():
        encontrado = None
        for alias in posibles:
            alias_n = _norm(alias)
            for i, h in enumerate(headers_norm):
                if i in excluir.values():   # no reutilizar índice ya asignado
                    continue
                if alias_n in h:
                    encontrado = i
                    break
            if encontrado is not None:
                break
        resultado[campo] = encontrado

    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# CARGA DESDE EL EXCEL DE IMPORTACIÓN  (misma hoja única que usa formacion.py)
# ─────────────────────────────────────────────────────────────────────────────

def _cargar_excel_unico(origen: str | bytes | io.BytesIO) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Lee el Excel de una sola hoja que acepta tu formulario.
    Columnas esperadas (igual que en formacion.html):
        Nombre del curso | Nombre | Progreso (%) | Exámenes | Fecha Inicio | Fecha Fin | Teléfono
    """
    import openpyxl

    if isinstance(origen, str):
        wb = openpyxl.load_workbook(origen, data_only=True)
    else:
        wb = openpyxl.load_workbook(
            io.BytesIO(origen) if isinstance(origen, bytes) else origen,
            data_only=True
        )

    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El Excel está vacío.")

    headers = [str(c) if c is not None else "" for c in filas[0]]

    # Detectar curso primero para excluirlo al buscar nombre
    idx_curso = _detectar_columnas(headers, {"nombre_curso": _ALIAS_ACADEMICOS["nombre_curso"]})
    excluir   = {k: v for k, v in idx_curso.items() if v is not None}
    idx_resto = _detectar_columnas(headers, {
        k: v for k, v in _ALIAS_ACADEMICOS.items() if k != "nombre_curso"
    }, excluir=excluir)
    idx_a = {**idx_curso, **idx_resto}
    idx_c = _detectar_columnas(headers, _ALIAS_CONTACTO)

    if idx_a.get("nombre_alumno") is None:
        raise ValueError(
            f"Columna 'Nombre' no encontrada. Cabeceras detectadas: {headers}"
        )

    registros_acad, registros_cont = [], []

    for fila in filas[1:]:
        if not any(fila):
            continue

        def get(campo, idx_map):
            i = idx_map.get(campo)
            return fila[i] if i is not None and i < len(fila) else None

        nombre = str(get("nombre_alumno", idx_a)).strip() if get("nombre_alumno", idx_a) else None
        if not nombre or nombre.lower() in ("none", "nan", ""):
            continue

        registros_acad.append({
            "nombre_curso":     str(get("nombre_curso", idx_a)).strip() if get("nombre_curso", idx_a) else None,
            "nombre_alumno":    nombre,
            "progreso_general": _safe_float(get("progreso_general", idx_a)),
            "examenes":         _safe_int(get("examenes", idx_a)),
        })

        tel = _limpiar_telefono(get("telefono", idx_c))
        if tel:
            registros_cont.append({"nombre_alumno": nombre, "telefono": tel})

    return pd.DataFrame(registros_acad), pd.DataFrame(registros_cont)


# ─────────────────────────────────────────────────────────────────────────────
# CARGA DESDE formacion.db
# ─────────────────────────────────────────────────────────────────────────────

def _cargar_desde_db(
    db_path: str = "formacion.db",
    tutor_id: int | None = None,
    incluir_archivados: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Extrae las dos fuentes de la tabla `alumnos` de formacion.db."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    filtros = []
    params: list[Any] = []

    if not incluir_archivados:
        filtros.append("(archivado IS NULL OR archivado = 0)")
    if tutor_id is not None:
        filtros.append("tutor_id = ?")
        params.append(tutor_id)

    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""

    df_acad = pd.read_sql_query(
        f"""
        SELECT
            curso    AS nombre_curso,
            nombre   AS nombre_alumno,
            progreso AS progreso_general,
            examenes AS examenes
        FROM alumnos {where}
        """,
        conn, params=params
    )

    # Fuente de contacto: misma tabla, solo filas con teléfono válido
    and_or_where = ("AND" if where else "WHERE")
    df_cont = pd.read_sql_query(
        f"""
        SELECT
            nombre   AS nombre_alumno,
            telefono AS telefono
        FROM alumnos
        {where} {and_or_where} telefono IS NOT NULL AND telefono != ''
        """,
        conn, params=params
    )

    conn.close()
    return df_acad, df_cont


# ─────────────────────────────────────────────────────────────────────────────
# MOTOR DE CONSOLIDACIÓN  (outer join por nombre normalizado)
# ─────────────────────────────────────────────────────────────────────────────

def consolidar(
    fuente_academica: list[dict] | pd.DataFrame,
    fuente_contacto:  list[dict] | pd.DataFrame,
    *,
    omitir_sin_telefono:  bool = False,
    omitir_sin_academico: bool = False,
) -> pd.DataFrame:
    """
    Cruza fuente académica y fuente de contacto por nombre normalizado.

    Parámetros
    ----------
    omitir_sin_telefono  : True → solo alumnos que tienen teléfono
    omitir_sin_academico : True → descarta contactos sin registro académico (left join)

    Resultado
    ---------
    DataFrame con columnas:
        nombre_curso | nombre_alumno | telefono | examenes | progreso_general
    Alumnos presentes solo en una fuente se conservan (outer join por defecto)
    con NaN en los campos que falten.
    """
    df_a = pd.DataFrame(fuente_academica) if isinstance(fuente_academica, list) else fuente_academica.copy()
    df_c = pd.DataFrame(fuente_contacto)  if isinstance(fuente_contacto,  list) else fuente_contacto.copy()

    # Garantizar columnas mínimas
    for col in ("nombre_curso", "nombre_alumno", "progreso_general", "examenes"):
        if col not in df_a.columns:
            df_a[col] = None
    for col in ("nombre_alumno", "telefono"):
        if col not in df_c.columns:
            df_c[col] = None

    # Clave de cruce: mismo algoritmo que norm() de formacion.py
    df_a["_clave"] = df_a["nombre_alumno"].apply(_norm)
    df_c["_clave"] = df_c["nombre_alumno"].apply(_norm)

    # Un teléfono por alumno (el primero no nulo)
    df_c = (
        df_c[df_c["telefono"].notna() & (df_c["telefono"] != "")]
        .drop_duplicates(subset="_clave", keep="first")
        [["_clave", "telefono"]]
    )

    how = "left" if omitir_sin_academico else "outer"
    merged = pd.merge(df_a, df_c, on="_clave", how=how)

    if omitir_sin_telefono:
        merged = merged[merged["telefono"].notna()]

    merged["examenes"]         = pd.to_numeric(merged["examenes"],         errors="coerce").fillna(0).astype(int)
    merged["progreso_general"] = pd.to_numeric(merged["progreso_general"],  errors="coerce").fillna(0.0).round(1)

    resultado = (
        merged[["nombre_curso", "nombre_alumno", "telefono", "examenes", "progreso_general"]]
        .sort_values(["nombre_curso", "nombre_alumno"], na_position="last")
        .reset_index(drop=True)
    )

    _imprimir_resumen(resultado)
    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# PUNTOS DE ENTRADA PÚBLICOS
# ─────────────────────────────────────────────────────────────────────────────

def consolidar_desde_db(
    db_path: str = "formacion.db",
    tutor_id: int | None = None,
    incluir_archivados: bool = False,
    **kwargs,
) -> pd.DataFrame:
    """
    Carga directamente desde formacion.db y devuelve el DataFrame consolidado.

    Uso típico dentro de formacion.py (Flask):

        from consolidar_alumnos import consolidar_desde_db

        @formacion_bp.route("/formacion/consolidado")
        @login_required
        def ver_consolidado():
            df = consolidar_desde_db(tutor_id=session["user_id"])
            return df.to_json(orient="records", force_ascii=False)
    """
    df_acad, df_cont = _cargar_desde_db(db_path, tutor_id, incluir_archivados)
    return consolidar(df_acad, df_cont, **kwargs)


def consolidar_desde_excel(
    ruta_excel: str | bytes | io.BytesIO,
    **kwargs,
) -> pd.DataFrame:
    """
    Carga el mismo Excel de una hoja que acepta tu formulario y devuelve
    el DataFrame consolidado sin tocar la BD.

    Columnas detectadas automáticamente (igual que en tu importación):
        Nombre del curso | Nombre | Progreso (%) | Exámenes | Teléfono ...

    Uso desde una ruta Flask:

        from consolidar_alumnos import consolidar_desde_excel

        @formacion_bp.route("/formacion/previsualizar", methods=["POST"])
        @login_required
        def previsualizar():
            df = consolidar_desde_excel(request.files["excel"].read())
            return df.to_json(orient="records", force_ascii=False)
    """
    df_acad, df_cont = _cargar_excel_unico(ruta_excel)
    return consolidar(df_acad, df_cont, **kwargs)


# ─────────────────────────────────────────────────────────────────────────────
# RESUMEN DE COBERTURA
# ─────────────────────────────────────────────────────────────────────────────

def _imprimir_resumen(df: pd.DataFrame) -> None:
    total = len(df)
    if total == 0:
        print("⚠️  El DataFrame consolidado está vacío.")
        return
    con_tel    = df["telefono"].notna().sum()
    sin_tel    = total - con_tel
    con_curso  = df["nombre_curso"].notna().sum()
    sin_curso  = total - con_curso
    superan_75 = (df["progreso_general"] >= 75).sum()
    print(
        f"\n📊 Consolidación completada:"
        f"\n   Total alumnos      : {total}"
        f"\n   Con teléfono       : {con_tel}  ({con_tel/total*100:.1f}%)"
        f"\n   Sin teléfono       : {sin_tel}  ({sin_tel/total*100:.1f}%)"
        f"\n   Con curso asignado : {con_curso}"
        f"\n   Sin curso asignado : {sin_curso}"
        f"\n   Superan 75%%       : {superan_75}  ({superan_75/total*100:.1f}%)\n"
    )


# ─────────────────────────────────────────────────────────────────────────────
# DEMO  (python consolidar_alumnos.py)
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import os

    if os.path.exists("formacion.db"):
        print("✅ Usando formacion.db real\n")
        df = consolidar_desde_db("formacion.db")
    else:
        print("⚠️  formacion.db no encontrada — usando datos de ejemplo\n")

        # Mismo formato que tu Excel: Nombre del curso | Nombre | Progreso | Exámenes
        fuente_academica = [
            {"nombre_curso": "Excel Avanzado", "nombre_alumno": "Ana García",    "progreso_general": 85.0, "examenes": 3},
            {"nombre_curso": "Excel Avanzado", "nombre_alumno": "Luis Martínez", "progreso_general": 60.0, "examenes": 2},
            {"nombre_curso": "Python Básico",  "nombre_alumno": "María López",   "progreso_general": 92.5, "examenes": 5},
            {"nombre_curso": "Python Básico",  "nombre_alumno": "Carlos Ruiz",   "progreso_general": 30.0, "examenes": 1},
            {"nombre_curso": "Excel Avanzado", "nombre_alumno": "Pedro Sanz",    "progreso_general": 10.0, "examenes": 0},
        ]
        # Columna Teléfono del mismo Excel (puede venir con variaciones de nombre)
        fuente_contacto = [
            {"nombre_alumno": "Ana García",    "telefono": "+34612345678"},
            {"nombre_alumno": "luis martinez", "telefono": "+34698765432"},  # minúsculas → OK
            {"nombre_alumno": "María López",   "telefono": "+34655111222"},
            # Carlos Ruiz y Pedro Sanz sin teléfono → aparecen con telefono=NaN
        ]
        df = consolidar(fuente_academica, fuente_contacto)

    print("─" * 65)
    print(df.to_string(index=False))

    print("\n📋 Resumen por curso:")
    resumen = (
        df.groupby("nombre_curso", dropna=False)
        .agg(
            alumnos        = ("nombre_alumno",    "count"),
            superan_75     = ("progreso_general",  lambda x: (x >= 75).sum()),
            prog_promedio  = ("progreso_general",  "mean"),
            total_examenes = ("examenes",          "sum"),
            con_telefono   = ("telefono",          lambda x: x.notna().sum()),
        )
        .round(1)
    )
    print(resumen.to_string())
