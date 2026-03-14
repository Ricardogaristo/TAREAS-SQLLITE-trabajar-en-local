"""
formacion.py  —  Módulo de Formación para el Gestor de Tareas
Base de datos independiente: formacion.db
"""

from flask import Blueprint, render_template, request, redirect, session, url_for, jsonify
from functools import wraps
from datetime import datetime
import openpyxl
import io
import os
import unicodedata
from db_mysql import get_form_conn, column_exists, get_db_name
from consolidar_alumnos import consolidar_desde_db, consolidar_desde_excel
from ia_formacion import (analizar_alumno, generar_mensaje_wa,
                           chatbot_tutor, limpiar_chat, predecir_riesgo_curso,
                           resumen_semanal, ranking_riesgo, mensajes_wa_masivos,
                           comparativa_cursos, sugerencias_hoy, importar_telefonos_excel)

# ── Blueprint ──────────────────────────────────────────────────────────────────
formacion_bp = Blueprint(
    "formacion", __name__,
    template_folder="templates"
)

FORM_DB = "formacion.db"

# ── Decorador de login ─────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ── Inicialización de tablas ───────────────────────────────────────────────────
def inicializar_formacion():
    conn = get_form_conn()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS alumnos (
            id              INT AUTO_INCREMENT PRIMARY KEY,
            curso           TEXT,
            nombre          TEXT NOT NULL,
            progreso        DOUBLE DEFAULT 0,
            examenes        VARCHAR(20) DEFAULT '0/0/0',
            fecha_inicio    VARCHAR(30),
            fecha_fin       VARCHAR(30),
            supera_75       INT DEFAULT 0,
            telefono        VARCHAR(50),
            tutor_id        INT,
            created_at      DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historial_snapshots (
            id           INT AUTO_INCREMENT PRIMARY KEY,
            tutor_id     INT,
            fecha        VARCHAR(20),
            label        TEXT,
            total        INT,
            superan_75   INT,
            pct_exito    DOUBLE,
            avg_progreso DOUBLE,
            created_at   DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historial_automatico (
            id             INT AUTO_INCREMENT PRIMARY KEY,
            tutor_id       INT,
            fecha          VARCHAR(20),
            evento         TEXT,
            total_alumnos  INT,
            total_cursos   INT,
            created_at     DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS alarmas_completadas (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            tutor_id    INT NOT NULL,
            clave       VARCHAR(150) NOT NULL,
            fecha_dia   VARCHAR(20) NOT NULL,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY unique_alarma (tutor_id, clave, fecha_dia)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS progreso_historial (
            id             INT AUTO_INCREMENT PRIMARY KEY,
            alumno_id      INT NOT NULL,
            tutor_id       INT NOT NULL,
            fecha_import   VARCHAR(20) NOT NULL,
            progreso       DOUBLE DEFAULT 0,
            examenes       INT DEFAULT 0,
            delta_progreso DOUBLE DEFAULT 0,
            avanzo         INT DEFAULT 0,
            FOREIGN KEY (alumno_id) REFERENCES alumnos(id) ON DELETE CASCADE
        )
    """)

    # Migraciones seguras — añade columnas si no existen
    for col, ddl in [
        ("curso",              "TEXT"),
        ("telefono",           "VARCHAR(50)"),
        ("supera_75",          "INT DEFAULT 0"),
        ("tutor_id",           "INT"),
        ("created_at",         "DATETIME DEFAULT CURRENT_TIMESTAMP"),
        ("archivado",          "INT DEFAULT 0"),
        ("archivado_at",       "VARCHAR(30)"),
        ("ultima_importacion", "VARCHAR(20)"),
        ("delta_progreso",     "DOUBLE DEFAULT 0"),
        ("avanzo",             "INT DEFAULT 0"),
        ("gestionado",         "INT DEFAULT 0"),
        ("tipo_gestion",       "VARCHAR(50)"),
        ("comentario",         "TEXT"),
        ("fecha_gestion",      "VARCHAR(30)"),
        ("no_llamar",          "TINYINT(1) DEFAULT 0"),
    ]:
        if not column_exists(cursor, 'alumnos', col):
            cursor.execute(f"ALTER TABLE alumnos ADD COLUMN {col} {ddl}")

    # Migrar examenes de INT a VARCHAR(20) si todavía es tipo numérico
    try:
        cursor.execute("ALTER TABLE alumnos MODIFY COLUMN examenes VARCHAR(20) DEFAULT '0/0/0'")
        cursor.execute("ALTER TABLE progreso_historial MODIFY COLUMN examenes VARCHAR(20) DEFAULT '0/0/0'")
    except Exception:
        pass  # Ya es VARCHAR o BD no soporta MODIFY

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS observaciones_alumno (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            alumno_id   INT NOT NULL,
            tutor_id    INT NOT NULL,
            texto       TEXT NOT NULL,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (alumno_id) REFERENCES alumnos(id) ON DELETE CASCADE
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS observaciones_alumno (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            alumno_id   INT NOT NULL,
            tutor_id    INT NOT NULL,
            texto       TEXT NOT NULL,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (alumno_id) REFERENCES alumnos(id) ON DELETE CASCADE
        )
    """)
    conn.commit()
    conn.close()
    print("✅ MySQL formacion inicializada correctamente.")

# ── Helpers ────────────────────────────────────────────────────────────────────
def _registrar_evento_historico(tutor_id, evento, conn):
    """Registra automáticamente un evento en el historial cada vez que cambia el estado."""
    row = conn.execute(
        "SELECT COUNT(*) as total, COUNT(DISTINCT curso) as cursos FROM alumnos WHERE tutor_id=?",
        (tutor_id,)
    ).fetchone()
    total_alumnos = row["total"] if row else 0
    total_cursos  = row["cursos"] if row else 0
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn.execute("""
        INSERT INTO historial_automatico (tutor_id, fecha, evento, total_alumnos, total_cursos)
        VALUES (?, ?, ?, ?, ?)
    """, (tutor_id, fecha, evento, total_alumnos, total_cursos))

def _safe_float(val):
    """Convierte a float; devuelve 0.0 si no es posible."""
    try:
        return float(str(val).replace("%", "").replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0

def _safe_int(val):
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return 0

def _fmt_examenes(val):
    """Garantiza que el valor de examenes esté en formato R/S/T.
    Acepta int (legacy) o string 'R/S/T'."""
    if val is None:
        return "0/0/0"
    s = str(val).strip()
    if "/" in s:
        parts = s.split("/")
        if len(parts) == 3:
            return s
    # Legacy: era un entero
    try:
        n = int(float(s))
        return f"{n}/0/0"
    except Exception:
        return "0/0/0"

def _parse_examenes(val):
    """Devuelve (realizados, superados, totales) como ints."""
    fmt = _fmt_examenes(val)
    parts = fmt.split("/")
    try:
        return int(parts[0]), int(parts[1]), int(parts[2])
    except Exception:
        return 0, 0, 0


def _safe_date(val):
    """Devuelve la fecha como string 'YYYY-MM-DD' o None."""
    if val is None:
        return None
    from datetime import datetime, date
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s if s else None


def _deduplicar_alumnos(tutor_id, conn):
    """
    Elimina duplicados (mismo nombre + mismo curso) para un tutor.
    De cada grupo duplicado conserva el registro con mayor progreso
    (o mayor id si hay empate). Los historial_progreso de los descartados
    se reasignan al registro conservado.
    Devuelve el número de duplicados eliminados.
    """
    import unicodedata as _ud

    def _n(s):
        return ''.join(c for c in _ud.normalize('NFD', str(s or ''))
                       if _ud.category(c) != 'Mn').lower().strip()

    # Todos los alumnos activos del tutor
    rows = conn.execute(
        "SELECT id, nombre, curso, progreso, examenes FROM alumnos "
        "WHERE tutor_id=? AND (archivado IS NULL OR archivado=0)",
        (tutor_id,)
    ).fetchall()

    # Agrupar por (nombre_norm, curso_norm)
    grupos = {}
    for r in rows:
        key = (_n(r["nombre"]), _n(r["curso"] or ""))
        grupos.setdefault(key, []).append(dict(r))

    eliminados = 0
    for key, grupo in grupos.items():
        if len(grupo) <= 1:
            continue
        # Conservar el de mayor progreso; en empate, el de mayor id
        grupo.sort(key=lambda x: (float(x["progreso"] or 0), x["id"]), reverse=True)
        keeper   = grupo[0]
        to_del   = grupo[1:]
        del_ids  = [x["id"] for x in to_del]

        # Reasignar historial de los duplicados al keeper
        for dup_id in del_ids:
            conn.execute(
                "UPDATE progreso_historial SET alumno_id=? WHERE alumno_id=?",
                (keeper["id"], dup_id)
            )

        # Borrar duplicados
        conn.execute(
            "DELETE FROM alumnos WHERE id IN ({})".format(
                ",".join("?" * len(del_ids))
            ),
            del_ids
        )
        eliminados += len(del_ids)

    return eliminados


# ── Motor de alarmas ──────────────────────────────────────────────────────────
def _generar_alarmas(tutor_id):
    """
    Genera la lista completa de alarmas del día para el tutor.
    Devuelve lista de dicts con: clave, tipo, prioridad, titulo, descripcion,
    alumno_id, alumno_nombre, curso, telefono, dias_restantes, accion_wa.
    """
    from datetime import date as _date_cls
    import urllib.parse

    hoy = _date_cls.today()
    conn = get_form_conn()
    alumnos = [dict(a) for a in conn.execute(
        "SELECT * FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0) ORDER BY fecha_fin ASC, progreso ASC",
        (tutor_id,)
    ).fetchall()]
    conn.close()

    alarmas = []
    cursos_foro_inicio = set()   # para no duplicar foros por curso
    cursos_foro_cierre = set()

    def _dias_restantes(fecha_fin_str):
        if not fecha_fin_str:
            return None
        try:
            return (_date_cls.fromisoformat(str(fecha_fin_str)[:10]) - hoy).days
        except Exception:
            return None

    def _dias_desde_inicio(fecha_inicio_str):
        if not fecha_inicio_str:
            return None
        try:
            return (hoy - _date_cls.fromisoformat(str(fecha_inicio_str)[:10])).days
        except Exception:
            return None

    def _wa_link(telefono, mensaje):
        if not telefono:
            return None
        tel = str(telefono).strip().replace(" ", "").replace("-", "").replace("+", "")
        return f"https://wa.me/{tel}?text={urllib.parse.quote(mensaje)}"

    for a in alumnos:
        alumno_id   = a["id"]
        nombre      = a.get("nombre", "")
        curso       = a.get("curso") or "Sin curso"
        progreso    = float(a.get("progreso") or 0)
        supera_75   = int(a.get("supera_75") or 0)
        telefono    = a.get("telefono") or ""
        fecha_fin   = a.get("fecha_fin")
        fecha_inicio= a.get("fecha_inicio")
        dias_r      = _dias_restantes(fecha_fin)
        dias_i      = _dias_desde_inicio(fecha_inicio)

        # ─── 1. CURSO VENCIDO Y NO APROBADO ──────────────────────────────────
        if dias_r is not None and dias_r < 0 and not supera_75:
            msg = (f"Hola {nombre} 👋, te contactamos porque tu curso *{curso}* ha finalizado "
                   f"y tu progreso está en {progreso:.0f}%. ¿Podemos ayudarte a completarlo? ¡Estamos aquí para acompañarte!")
            alarmas.append({
                "clave"         : f"vencido:{alumno_id}",
                "tipo"          : "vencido_bajo",
                "prioridad"     : 1,
                "emoji"         : "🔴",
                "titulo"        : "Curso vencido sin aprobar",
                "descripcion"   : f"El curso finalizó hace {abs(dias_r)} día{'s' if abs(dias_r)!=1 else ''} y el alumno tiene {progreso:.0f}% de progreso.",
                "accion"        : "Contactar urgente por WhatsApp y evaluar prórroga.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })

        # ─── 2. PROGRESO CRÍTICO CON POCO TIEMPO ────────────────────────────
        if dias_r is not None and 0 <= dias_r <= 20 and progreso < 25 and not supera_75:
            msg = (f"Hola {nombre} 👋, notamos que tu progreso en *{curso}* es de solo {progreso:.0f}% "
                   f"y quedan {dias_r} días para finalizar. ¡Te acompañamos para que puedas alcanzar el objetivo! ¿Cuándo podemos hablar?")
            alarmas.append({
                "clave"         : f"critico:{alumno_id}",
                "tipo"          : "progreso_critico",
                "prioridad"     : 1,
                "emoji"         : "🔴",
                "titulo"        : "Alumno en riesgo crítico",
                "descripcion"   : f"Progreso {progreso:.0f}% con solo {dias_r} días restantes. Riesgo alto de no aprobar.",
                "accion"        : "Intervención inmediata. Llamar o enviar WA con plan de acción.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })

        # ─── 3. FECHA FIN EN 7 DÍAS ──────────────────────────────────────────
        elif dias_r is not None and 1 <= dias_r <= 7 and not supera_75:
            msg = (f"Hola {nombre} 👋, te recordamos que tu curso *{curso}* finaliza en {dias_r} día{'s' if dias_r!=1 else ''}. "
                   f"Tu progreso actual es {progreso:.0f}%. ¡Ánimo, todavía estás a tiempo! 💪")
            alarmas.append({
                "clave"         : f"fin7:{alumno_id}",
                "tipo"          : "fin_7dias",
                "prioridad"     : 1,
                "emoji"         : "🔴",
                "titulo"        : f"Cierre en {dias_r} día{'s' if dias_r!=1 else ''} — sin aprobar",
                "descripcion"   : f"El curso termina el {fecha_fin}. Progreso: {progreso:.0f}%.",
                "accion"        : f"Enviar WA de recordatorio urgente de cierre.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })

        # ─── 4. FECHA FIN EN 8–14 DÍAS ───────────────────────────────────────
        elif dias_r is not None and 8 <= dias_r <= 14 and not supera_75:
            msg = (f"Hola {nombre} 👋, quedan {dias_r} días para que finalice tu curso *{curso}*. "
                   f"Llevas un progreso de {progreso:.0f}%. ¡Te animamos a avanzar para alcanzar el objetivo! 🎓")
            alarmas.append({
                "clave"         : f"fin14:{alumno_id}",
                "tipo"          : "fin_14dias",
                "prioridad"     : 2,
                "emoji"         : "🟡",
                "titulo"        : f"Cierre en {dias_r} días — recordatorio",
                "descripcion"   : f"Quedan 2 semanas. Progreso: {progreso:.0f}%. Buen momento para hacer seguimiento.",
                "accion"        : "Enviar WA motivacional con recordatorio de fechas.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })

        # ─── 5. FECHA FIN EN 15–30 DÍAS Y PROGRESO BAJO ─────────────────────
        elif dias_r is not None and 15 <= dias_r <= 30 and progreso < 50 and not supera_75:
            msg = (f"Hola {nombre} 👋, ¿cómo vas con tu curso *{curso}*? Tienes {dias_r} días para completarlo "
                   f"y tu progreso es {progreso:.0f}%. Planifiquemos juntos para que llegues a la meta 🎯")
            alarmas.append({
                "clave"         : f"fin30:{alumno_id}",
                "tipo"          : "fin_30dias",
                "prioridad"     : 2,
                "emoji"         : "🟡",
                "titulo"        : f"Progreso bajo con cierre en {dias_r} días",
                "descripcion"   : f"Progreso {progreso:.0f}% — menos del 50% con menos de 30 días disponibles.",
                "accion"        : "Planificar seguimiento y enviar WA con agenda de trabajo.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })

        # ─── 6. INICIO HOY ───────────────────────────────────────────────────
        if dias_i is not None and dias_i == 0:
            msg = (f"¡Bienvenido/a {nombre}! 🎉 Hoy comenzás tu curso *{curso}*. "
                   f"Estamos muy contentos de acompañarte en este proceso de aprendizaje. "
                   f"¡Cualquier consulta, aquí estamos! 😊")
            alarmas.append({
                "clave"         : f"inicio_hoy:{alumno_id}",
                "tipo"          : "inicio_hoy",
                "prioridad"     : 2,
                "emoji"         : "🟢",
                "titulo"        : "Inicio de curso HOY",
                "descripcion"   : f"El alumno empieza hoy el curso {curso}.",
                "accion"        : "Enviar WA de bienvenida y verificar acceso al aula virtual.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })
            # Foro de inicio — una alarma por curso, no por alumno
            if curso not in cursos_foro_inicio:
                cursos_foro_inicio.add(curso)
                alarmas.append({
                    "clave"         : f"foro_inicio:{curso}",
                    "tipo"          : "foro_inicio",
                    "prioridad"     : 2,
                    "emoji"         : "🟢",
                    "titulo"        : f"Publicar foro de inicio — {curso}",
                    "descripcion"   : f"El curso {curso} arranca hoy. Publicar el foro de presentación e inicio de la primera unidad.",
                    "accion"        : "Abrir hilo de presentación y foro de inicio de unidad en el aula virtual.",
                    "alumno_id"     : None,
                    "alumno_nombre" : f"Curso completo: {curso}",
                    "curso"         : curso,
                    "telefono"      : "",
                    "dias_restantes": None,
                    "wa_link"       : None,
                })

        # ─── 7. INICIO HACE 1–3 DÍAS ─────────────────────────────────────────
        elif dias_i is not None and 1 <= dias_i <= 3:
            alarmas.append({
                "clave"         : f"inicio_reciente:{alumno_id}",
                "tipo"          : "inicio_reciente",
                "prioridad"     : 3,
                "emoji"         : "🔵",
                "titulo"        : f"Inicio reciente ({dias_i} día{'s' if dias_i!=1 else ''} atrás)",
                "descripcion"   : f"El alumno inició el {fecha_inicio}. Verificar que ingresó correctamente al aula.",
                "accion"        : "Confirmar acceso al aula virtual y progreso inicial.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : None,
            })

        # ─── 8. FORO DE CIERRE DE UNIDAD ─────────────────────────────────────
        if dias_r is not None and 3 <= dias_r <= 7 and curso not in cursos_foro_cierre:
            cursos_foro_cierre.add(curso)
            alarmas.append({
                "clave"         : f"foro_cierre:{curso}",
                "tipo"          : "foro_cierre",
                "prioridad"     : 2,
                "emoji"         : "🟡",
                "titulo"        : f"Publicar foro de cierre — {curso}",
                "descripcion"   : f"El curso {curso} cierra en {dias_r} días. Publicar foro de cierre de unidad y evaluación final.",
                "accion"        : "Abrir foro de cierre, reflexión final y recordatorio de evaluación.",
                "alumno_id"     : None,
                "alumno_nombre" : f"Curso completo: {curso}",
                "curso"         : curso,
                "telefono"      : "",
                "dias_restantes": dias_r,
                "wa_link"       : None,
            })

        # ─── 9. SIN TELÉFONO REGISTRADO ──────────────────────────────────────
        if not telefono and (
            (dias_r is not None and dias_r <= 30) or progreso < 50
        ):
            alarmas.append({
                "clave"         : f"sin_tel:{alumno_id}",
                "tipo"          : "sin_telefono",
                "prioridad"     : 3,
                "emoji"         : "🔵",
                "titulo"        : "Sin teléfono de contacto",
                "descripcion"   : f"El alumno no tiene teléfono registrado y requiere seguimiento (progreso {progreso:.0f}%).",
                "accion"        : "Registrar número de WhatsApp en la ficha del alumno.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : "",
                "dias_restantes": dias_r,
                "wa_link"       : None,
            })

        # ─── 10. PROGRESO MENOR AL 40% ───────────────────────────────────────
        if progreso < 40 and not supera_75 and not (
            # Evitar duplicar con las alarmas críticas ya generadas
            (dias_r is not None and dias_r <= 20 and progreso < 25) or
            (dias_r is not None and dias_r < 0)
        ):
            msg = (f"Hola {nombre} 👋, notamos que tu progreso en *{curso}* es de {progreso:.0f}%, "
                   f"que está por debajo del 40% esperado. "
                   f"¿Hay algo en lo que podamos ayudarte para avanzar? "
                   f"¡Estamos aquí para acompañarte! 💪")
            alarmas.append({
                "clave"         : f"bajo40:{alumno_id}",
                "tipo"          : "progreso_bajo_40",
                "prioridad"     : 2,
                "emoji"         : "🟠",
                "titulo"        : f"Progreso bajo el 40% — {progreso:.0f}%",
                "descripcion"   : f"El alumno lleva {progreso:.0f}% de avance, por debajo del umbral mínimo del 40%."
                                  + (f" Quedan {dias_r} día{'s' if dias_r != 1 else ''}." if dias_r is not None and dias_r > 0 else ""),
                "accion"        : "Contactar al alumno para identificar obstáculos y reforzar el acompañamiento.",
                "alumno_id"     : alumno_id,
                "alumno_nombre" : nombre,
                "curso"         : curso,
                "telefono"      : telefono,
                "dias_restantes": dias_r,
                "wa_link"       : _wa_link(telefono, msg),
            })

    # Ordenar por prioridad ascendente, luego días restantes
    alarmas.sort(key=lambda x: (
        x["prioridad"],
        x["dias_restantes"] if x["dias_restantes"] is not None else 9999
    ))
    return alarmas


def _get_completadas_hoy(tutor_id):
    """Devuelve el set de claves completadas hoy por este tutor."""
    from datetime import date as _d
    hoy = _d.today().isoformat()
    conn = get_form_conn()
    rows = conn.execute(
        "SELECT clave FROM alarmas_completadas WHERE tutor_id=? AND fecha_dia=?",
        (tutor_id, hoy)
    ).fetchall()
    conn.close()
    return {r["clave"] for r in rows}

# ── Ruta: listado de alumnos + carga de Excel ──────────────────────────────────
@formacion_bp.route("/formacion", methods=["GET", "POST"])
@login_required
def formacion():
    tutor_id = session.get("user_id")
    errores  = []
    exito    = None

    if request.method == "POST":
        arch = request.files.get("excel")

        if not arch or not arch.filename.endswith((".xlsx", ".xls")):
            errores.append("Sube un archivo Excel (.xlsx o .xls).")

        if not errores:
            try:
                fecha_import = datetime.now().strftime("%Y-%m-%d")

                # ── normalización ──────────────────────────────────────────
                def norm(s):
                    return ''.join(c for c in unicodedata.normalize('NFD', str(s))
                                   if unicodedata.category(c) != 'Mn').lower().strip()

                def get_col(hn, posibles, excluir=None):
                    for p in posibles:
                        pn = norm(p)
                        for i, h in enumerate(hn):
                            if pn in h and i != excluir:
                                return i
                    return None

                # ── Único Excel: todas las columnas en una sola hoja ─────
                wb2  = openpyxl.load_workbook(io.BytesIO(arch.read()), data_only=True)
                ws2  = wb2.active
                hn2  = [norm(c.value) if c.value else "" for c in next(ws2.iter_rows(min_row=1, max_row=1))]
                i2_c  = get_col(hn2, ["del curso","curso","materia","asignatura"])
                i2_n  = get_col(hn2, ["nombre","alumno","estudiante"], excluir=i2_c)
                i2_p  = get_col(hn2, ["progreso","avance","progress"])
                i2_e  = get_col(hn2, ["examen","exam","evaluac","prueba"])
                i2_fi = get_col(hn2, ["fecha inicio","fecha de inicio","f. inicio","inicio"])
                i2_ff = next((i for i,h in enumerate(hn2) if "fin" in h and "inicio" not in h), None)
                i2_t  = get_col(hn2, ["telefono","celular","phone","whatsapp","movil","tel"])
                if i2_n is None:
                    errores.append("Columna 'Nombre' no encontrada en el Excel.")
                    raise ValueError("headers")

                # Construir tel_map desde la misma hoja
                tel_map = {}
                if i2_t is not None:
                    for row in ws2.iter_rows(min_row=2, values_only=True):
                        if not any(row): continue
                        n = row[i2_n]
                        t = row[i2_t]
                        if n and t:
                            tel_map[norm(str(n))] = str(t).strip().replace(".0","").replace(" ","")
                print(f"📱 {len(tel_map)} teléfonos cargados")

                # ── UPSERT con historial ──────────────────────────────────
                conn = get_form_conn()
                curs = conn.cursor()

                # Primero limpiar duplicados ya existentes en la BD
                dup_eliminados = _deduplicar_alumnos(tutor_id, conn)
                if dup_eliminados:
                    print(f"🧹 {dup_eliminados} duplicados eliminados antes de importar")

                # Mapa de existentes: (nombre_norm, curso_norm) → {id, progreso, examenes}
                existentes = {}
                for ex in conn.execute(
                    "SELECT id, nombre, curso, progreso, examenes FROM alumnos "
                    "WHERE tutor_id=? AND (archivado IS NULL OR archivado=0)", (tutor_id,)
                ).fetchall():
                    key = (norm(ex["nombre"]), norm(ex["curso"] or ""))
                    existentes[key] = dict(ex)

                cnt_nuevo = cnt_update = cnt_avanza = cnt_retro = cnt_igual = cnt_sin_tel = 0

                for row in ws2.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    nombre = str(row[i2_n]).strip() if row[i2_n] else None
                    if not nombre or nombre.lower() in ("none","nan",""): continue

                    progreso = _safe_float(row[i2_p])  if i2_p  is not None and i2_p  < len(row) else 0.0
                    examenes = _fmt_examenes(row[i2_e] if i2_e is not None and i2_e < len(row) else None)
                    f_inicio = _safe_date(row[i2_fi])  if i2_fi is not None and i2_fi < len(row) else None
                    f_fin    = _safe_date(row[i2_ff])  if i2_ff is not None and i2_ff < len(row) else None
                    curso    = str(row[i2_c]).strip()  if i2_c  is not None and i2_c  < len(row) and row[i2_c] else None
                    supera75 = 1 if progreso >= 75 else 0
                    telefono = tel_map.get(norm(nombre))
                    if not telefono: cnt_sin_tel += 1

                    key      = (norm(nombre), norm(curso or ""))
                    existing = existentes.get(key)

                    if existing:
                        alumno_id    = existing["id"]
                        old_progreso = float(existing["progreso"] or 0)
                        delta        = round(progreso - old_progreso, 2)
                        avanzo       = 1 if delta > 0 else (-1 if delta < 0 else 0)
                        if avanzo ==  1: cnt_avanza += 1
                        elif avanzo == -1: cnt_retro += 1
                        else: cnt_igual += 1
                        curs.execute("""
                            UPDATE alumnos
                            SET progreso=?, examenes=?, fecha_inicio=?, fecha_fin=?,
                                supera_75=?, telefono=COALESCE(?,telefono),
                                ultima_importacion=?, delta_progreso=?, avanzo=?
                            WHERE id=?
                        """, (progreso, examenes, f_inicio, f_fin, supera75,
                              telefono, fecha_import, delta, avanzo, alumno_id))
                        cnt_update += 1
                    else:
                        delta = 0.0; avanzo = 0
                        curs.execute("""
                            INSERT INTO alumnos
                                (curso, nombre, progreso, examenes, fecha_inicio, fecha_fin,
                                 supera_75, telefono, tutor_id, ultima_importacion, delta_progreso, avanzo)
                            VALUES (?,?,?,?,?,?,?,?,?,?,0,0)
                        """, (curso, nombre, progreso, examenes, f_inicio, f_fin,
                              supera75, telefono, tutor_id, fecha_import))
                        alumno_id = curs.lastrowid
                        cnt_nuevo += 1

                    # Registro historial de progreso
                    old_p = float(existing["progreso"] or 0) if existing else 0.0
                    curs.execute("""
                        INSERT INTO progreso_historial
                            (alumno_id, tutor_id, fecha_import, progreso, examenes, delta_progreso, avanzo)
                        VALUES (?,?,?,?,?,?,?)
                    """, (alumno_id, tutor_id, fecha_import, progreso, examenes,
                          round(progreso - old_p, 2) if existing else 0.0,
                          1 if (progreso - old_p) > 0 else (-1 if (progreso - old_p) < 0 else 0) if existing else 0))

                    # Actualizar mapa para evitar duplicados dentro del mismo Excel
                    existentes[key] = {"id": alumno_id, "nombre": nombre,
                                       "curso": curso, "progreso": progreso, "examenes": examenes}

                _registrar_evento_historico(
                    tutor_id,
                    f"Importación ({cnt_nuevo} nuevos, {cnt_update} actualizados)",
                    conn
                )
                conn.commit()
                conn.close()
                dup_msg = f" · {dup_eliminados} duplicados fusionados" if dup_eliminados else ""
                exito = (f"✅ {cnt_nuevo + cnt_update} alumnos procesados — "
                         f"{cnt_nuevo} nuevos · {cnt_update} actualizados "
                         f"({cnt_avanza} avanzaron, {cnt_retro} retrocedieron, {cnt_igual} sin cambio)"
                         f"{dup_msg} · {cnt_sin_tel} sin teléfono.")

            except ValueError:
                pass
            except Exception as e:
                errores.append(f"Error al procesar los archivos: {e}")

    # Cargar alumnos activos del tutor actual (excluye archivados)
    conn = get_form_conn()
    _alumnos_raw = conn.execute(
        "SELECT * FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0) ORDER BY progreso DESC, nombre ASC",
        (tutor_id,)
    ).fetchall()
    alumnos = []
    for _a in _alumnos_raw:
        _a = dict(_a)
        _a["examenes"]   = _fmt_examenes(_a.get("examenes"))
        _ep              = _parse_examenes(_a["examenes"])
        _a["ex_pending"] = _ep[2] > 0 and (_ep[0] / _ep[2]) < 0.75 and not _a.get("no_llamar")
        alumnos.append(_a)
    # Contar cursos archivados para el badge
    row_arch = conn.execute(
        "SELECT COUNT(DISTINCT curso) as n FROM alumnos WHERE tutor_id=? AND archivado=1",
        (tutor_id,)
    ).fetchone()
    archivados_count = row_arch["n"] if row_arch else 0

    # Número de importaciones distintas por curso (cuenta fechas únicas en progreso_historial)
    imp_rows = conn.execute(
        """SELECT a.curso, COUNT(DISTINCT ph.fecha_import) as n_imports
           FROM progreso_historial ph
           JOIN alumnos a ON ph.alumno_id = a.id
           WHERE ph.tutor_id=? AND (a.archivado IS NULL OR a.archivado=0)
           GROUP BY a.curso""",
        (tutor_id,)
    ).fetchall()
    importaciones_por_curso = {r["curso"]: r["n_imports"] for r in imp_rows}

    conn.close()

    # Contar alarmas pendientes del día para el badge
    alarmas_hoy      = _generar_alarmas(tutor_id)
    completadas_hoy  = _get_completadas_hoy(tutor_id)
    alarmas_pendientes = sum(1 for a in alarmas_hoy if a["clave"] not in completadas_hoy)

    return render_template("formacion.html", alumnos=alumnos, errores=errores, exito=exito,
                           alarmas_pendientes=alarmas_pendientes,
                           archivados_count=archivados_count,
                           importaciones_por_curso=importaciones_por_curso)

@formacion_bp.route("/formacion/deduplicar", methods=["POST"])
@login_required
def deduplicar_alumnos_route():
    """Ejecuta la deduplicación manualmente y redirige con mensaje."""
    tutor_id = session.get("user_id")
    conn = get_form_conn()
    n = _deduplicar_alumnos(tutor_id, conn)
    conn.commit()
    conn.close()
    return jsonify({"eliminados": n})


@formacion_bp.route("/formacion/historial_alumno/<int:alumno_id>")
@login_required
def historial_alumno(alumno_id):
    """JSON: devuelve el historial de progreso de un alumno."""
    tutor_id = session.get("user_id")
    conn = get_form_conn()
    rows = conn.execute(
        "SELECT fecha_import, progreso, examenes, delta_progreso, avanzo "
        "FROM progreso_historial WHERE alumno_id=? AND tutor_id=? ORDER BY fecha_import ASC",
        (alumno_id, tutor_id)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@formacion_bp.route("/formacion/borrar-curso", methods=["POST"])
@login_required
def borrar_curso():
    tutor_id = session.get("user_id")
    curso = request.form.get("curso", "").strip()
    if not curso:
        return redirect(url_for("formacion.formacion"))
    conn = get_form_conn()
    conn.execute("DELETE FROM alumnos WHERE curso=? AND tutor_id=?", (curso, tutor_id))
    _registrar_evento_historico(tutor_id, f"Borrado curso: {curso}", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))

# ── Ruta: toggle no_llamar (AJAX) ─────────────────────────────────────────────
@formacion_bp.route("/formacion/alumno/no_llamar/<int:alumno_id>", methods=["POST"])
@login_required
def toggle_no_llamar(alumno_id):
    datos = request.get_json(silent=True) or {}
    valor = 1 if datos.get("no_llamar") else 0
    conn  = get_form_conn()
    conn.execute(
        "UPDATE alumnos SET no_llamar=? WHERE id=? AND (tutor_id=? OR tutor_id IS NULL)",
        (valor, alumno_id, session["user_id"])
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "no_llamar": bool(valor)})


# ── Ruta: editar alumno (teléfono) ─────────────────────────────────────────────
@formacion_bp.route("/formacion/editar/<int:alumno_id>", methods=["POST"])
@login_required
def editar_alumno(alumno_id):
    telefono = request.form.get("telefono", "").strip()
    conn = get_form_conn()
    # Permite editar si el alumno pertenece al tutor O si tutor_id es NULL (migrado)
    conn.execute("""UPDATE alumnos SET telefono=?
                    WHERE id=? AND (tutor_id=? OR tutor_id IS NULL)""",
                 (telefono, alumno_id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: eliminar alumno ──────────────────────────────────────────────────────
@formacion_bp.route("/formacion/eliminar/<int:alumno_id>")
@login_required
def eliminar_alumno(alumno_id):
    conn = get_form_conn()
    conn.execute("DELETE FROM alumnos WHERE id=? AND tutor_id=?",
                 (alumno_id, session["user_id"]))
    _registrar_evento_historico(session["user_id"], "Eliminación de alumno", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: borrar TODOS los alumnos ────────────────────────────────────────────
@formacion_bp.route("/formacion/borrar_todos", methods=["POST"])
@login_required
def borrar_todos():
    conn = get_form_conn()
    conn.execute("DELETE FROM alumnos WHERE tutor_id=?", (session["user_id"],))
    _registrar_evento_historico(session["user_id"], "Borrado total", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: archivar curso completo ─────────────────────────────────────────────
@formacion_bp.route("/formacion/archivar_curso", methods=["POST"])
@login_required
def archivar_curso():
    tutor_id = session["user_id"]
    curso    = request.form.get("curso", "").strip()
    if not curso:
        return redirect(url_for("formacion.formacion"))

    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn  = get_form_conn()
    conn.execute("""
        UPDATE alumnos
        SET archivado=1, archivado_at=?
        WHERE tutor_id=? AND curso=? AND (archivado IS NULL OR archivado=0)
    """, (ahora, tutor_id, curso))
    _registrar_evento_historico(tutor_id, f"Curso archivado: {curso}", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion"))


# ── Ruta: restaurar curso archivado ───────────────────────────────────────────
@formacion_bp.route("/formacion/restaurar_curso", methods=["POST"])
@login_required
def restaurar_curso():
    tutor_id = session["user_id"]
    curso    = request.form.get("curso", "").strip()
    if not curso:
        return redirect(url_for("formacion.formacion_archivados"))

    conn = get_form_conn()
    conn.execute("""
        UPDATE alumnos
        SET archivado=0, archivado_at=NULL
        WHERE tutor_id=? AND curso=? AND archivado=1
    """, (tutor_id, curso))
    _registrar_evento_historico(tutor_id, f"Curso restaurado: {curso}", conn)
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_archivados"))


# ── Ruta: vista de archivados ──────────────────────────────────────────────────
@formacion_bp.route("/formacion/archivados")
@login_required
def formacion_archivados():
    tutor_id = session["user_id"]
    conn     = get_form_conn()

    alumnos_arch = [dict(a) for a in conn.execute(
        """SELECT * FROM alumnos
           WHERE tutor_id=? AND archivado=1
           ORDER BY archivado_at DESC, curso, nombre""",
        (tutor_id,)
    ).fetchall()]
    conn.close()

    # Normalizar
    for a in alumnos_arch:
        a["progreso"]  = float(a.get("progreso") or 0)
        a["examenes"]  = _fmt_examenes(a.get("examenes"))
        _ep = _parse_examenes(a["examenes"])
        a["ex_pending"] = _ep[2] > 0 and (_ep[0] / _ep[2]) < 0.75  # ratio realizados/totales < 75%
        a["supera_75"] = int(a.get("supera_75") or 0)
        a["curso"]     = a.get("curso") or "Sin curso"

    # Agrupar por curso
    from collections import defaultdict, OrderedDict
    grupos_raw = defaultdict(list)
    for a in alumnos_arch:
        grupos_raw[a["curso"]].append(a)

    # ── Fusionar alumnos duplicados dentro de cada curso ──────────────────────
    # Mismo nombre (normalizado) en el mismo curso → conservar el de mayor
    # progreso (o mayor id en empate) y calcular delta vs el anterior.
    def _norm_name(s):
        import unicodedata as _ud
        return ''.join(c for c in _ud.normalize('NFD', str(s or ''))
                       if _ud.category(c) != 'Mn').lower().strip()

    for curso_key in grupos_raw:
        lista = grupos_raw[curso_key]
        # Sub-agrupar por nombre normalizado
        por_nombre = defaultdict(list)
        for a in lista:
            por_nombre[_norm_name(a["nombre"])].append(a)

        fusionados = []
        for nombre_norm, duplicados in por_nombre.items():
            if len(duplicados) == 1:
                a = duplicados[0]
                a["progreso_anterior"] = None   # sin dato previo
                a["delta_arch"]        = None
                a["tendencia"]         = "nuevo"
                fusionados.append(a)
            else:
                # Ordenar por progreso desc, luego id desc → [0] es el más avanzado
                duplicados.sort(key=lambda x: (x["progreso"], x["id"]), reverse=True)
                keeper   = duplicados[0]
                # El "anterior" es el segundo mejor (o el más antiguo si hay muchos)
                otros    = duplicados[1:]
                prog_ant = max(x["progreso"] for x in otros)
                delta    = round(keeper["progreso"] - prog_ant, 1)
                if delta > 0:
                    tendencia = "subio"
                elif delta < 0:
                    tendencia = "bajo"
                else:
                    tendencia = "igual"
                keeper["progreso_anterior"] = prog_ant
                keeper["delta_arch"]        = delta
                keeper["tendencia"]         = tendencia
                fusionados.append(keeper)

        grupos_raw[curso_key] = fusionados

    # Ordenar grupos por fecha de archivo (más reciente primero)
    def _fecha_archivo_grupo(alumnos_lista):
        fechas = [a.get("archivado_at") or "" for a in alumnos_lista]
        return max(fechas) if fechas else ""

    grupos_ordenados = sorted(grupos_raw.items(),
                              key=lambda kv: _fecha_archivo_grupo(kv[1]),
                              reverse=True)

    # Calcular estadísticas por grupo (excluye alumnos con no_llamar=1)
    grupos = []
    for curso_nombre, lista in grupos_ordenados:
        total     = len(lista)
        contables = [a for a in lista if not a.get("no_llamar")]   # excluir no_llamar
        n_cont    = len(contables)
        superan   = sum(1 for a in contables if a["supera_75"] == 1)
        pct       = round(superan / n_cont * 100, 1) if n_cont else 0
        avg_prog  = round(sum(a["progreso"] for a in contables) / n_cont, 1) if n_cont else 0
        examenes  = sum(_parse_examenes(a["examenes"])[0] for a in lista)  # suma de realizados
        fecha_arch = (lista[0].get("archivado_at") or "")[:10]
        fecha_ini  = min((a.get("fecha_inicio") or "9999" for a in lista if a.get("fecha_inicio")), default="—")
        fecha_fin  = max((a.get("fecha_fin")    or "0000" for a in lista if a.get("fecha_fin")),    default="—")
        grupos.append({
            "curso"      : curso_nombre,
            "alumnos"    : lista,
            "total"      : total,
            "total_contable": n_cont,
            "superan"    : superan,
            "pct_exito"  : pct,
            "avg_progreso": avg_prog,
            "total_examenes": examenes,  # realizados totales del grupo
            "fecha_archivo": fecha_arch,
            "fecha_inicio" : fecha_ini if fecha_ini != "9999" else "—",
            "fecha_fin"    : fecha_fin if fecha_fin != "0000" else "—",
        })

    total_archivados = len(alumnos_arch)
    total_cursos_arch = len(grupos)

    return render_template(
        "formacion_archivados.html",
        grupos=grupos,
        total_archivados=total_archivados,
        total_cursos_arch=total_cursos_arch,
    )
@formacion_bp.route("/formacion/guardar_snapshot", methods=["POST"])
@login_required
def guardar_snapshot():
    tutor_id = session["user_id"]
    label    = request.form.get("label", "").strip() or datetime.now().strftime("%d/%m/%Y")
    fecha    = datetime.now().strftime("%Y-%m-%d")

    conn    = get_form_conn()
    alumnos = conn.execute(
        "SELECT * FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0)", (tutor_id,)
    ).fetchall()

    total        = len(alumnos)
    superan_75   = sum(1 for a in alumnos if a["supera_75"] == 1)
    pct_exito    = round(superan_75 / total * 100, 1) if total else 0
    avg_progreso = round(sum(a["progreso"] for a in alumnos) / total, 1) if total else 0

    conn.execute("""
        INSERT INTO historial_snapshots (tutor_id, fecha, label, total, superan_75, pct_exito, avg_progreso)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (tutor_id, fecha, label, total, superan_75, pct_exito, avg_progreso))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_dashboard"))


# ── Ruta: borrar snapshot ──────────────────────────────────────────────────────
@formacion_bp.route("/formacion/borrar_snapshot/<int:snap_id>")
@login_required
def borrar_snapshot(snap_id):
    conn = get_form_conn()
    conn.execute("DELETE FROM historial_snapshots WHERE id=? AND tutor_id=?",
                 (snap_id, session["user_id"]))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_dashboard"))


# ── Ruta: borrar todo el historial automático ─────────────────────────────────
@formacion_bp.route("/formacion/borrar_historial_auto", methods=["POST"])
@login_required
def borrar_historial_auto():
    conn = get_form_conn()
    conn.execute("DELETE FROM historial_automatico WHERE tutor_id=?", (session["user_id"],))
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_dashboard"))
@formacion_bp.route("/formacion/alarmas")
@login_required
def formacion_alarmas():
    tutor_id   = session["user_id"]
    alarmas    = _generar_alarmas(tutor_id)
    completadas = _get_completadas_hoy(tutor_id)

    # Separar pendientes y completadas
    for a in alarmas:
        a["completada"] = a["clave"] in completadas

    total      = len(alarmas)
    pendientes = sum(1 for a in alarmas if not a["completada"])
    hechas     = total - pendientes

    from datetime import date as _d
    hoy_str = _d.today().strftime("%A %d de %B de %Y")

    # ── Resumen de cursos para los KPI cards ──────────────────────
    conn2 = get_form_conn()
    alumnos_cursos = [dict(a) for a in conn2.execute(
        "SELECT curso, progreso, supera_75, fecha_inicio, fecha_fin, no_llamar FROM alumnos "
        "WHERE tutor_id=? AND (archivado IS NULL OR archivado=0)",
        (tutor_id,)
    ).fetchall()]
    conn2.close()

    from collections import defaultdict
    _cur = defaultdict(lambda: {"total":0,"superan":0,"prog":0,"f_ini":None,"f_fin":None})
    for a in alumnos_cursos:
        if a.get("no_llamar"): continue
        k = a["curso"] or "Sin curso"
        _cur[k]["total"]  += 1
        _cur[k]["superan"] += int(a.get("supera_75") or 0)
        _cur[k]["prog"]   += float(a.get("progreso") or 0)
        if a.get("fecha_inicio") and not _cur[k]["f_ini"]:
            _cur[k]["f_ini"] = a["fecha_inicio"]
        if a.get("fecha_fin") and not _cur[k]["f_fin"]:
            _cur[k]["f_fin"] = a["fecha_fin"]

    cursos_resumen = []
    for nombre_c, d in sorted(_cur.items()):
        n = d["total"]
        avg = round(d["prog"] / n, 1) if n else 0
        sup = d["superan"]
        pct_exito = round(sup / n * 100, 1) if n else 0
        cursos_resumen.append({
            "nombre":    nombre_c,
            "total":     n,
            "superan":   sup,
            "pct_exito": pct_exito,
            "avg_prog":  avg,
            "f_ini":     (d["f_ini"] or "")[:10],
            "f_fin":     (d["f_fin"] or "")[:10],
        })

    return render_template(
        "formacion_alarmas.html",
        alarmas=alarmas,
        total=total,
        pendientes=pendientes,
        hechas=hechas,
        hoy_str=hoy_str,
        cursos_resumen=cursos_resumen,
    )


@formacion_bp.route("/formacion/alarmas/completar", methods=["POST"])
@login_required
def alarma_completar():
    from datetime import date as _d
    tutor_id = session["user_id"]
    clave    = request.form.get("clave", "").strip()
    accion   = request.form.get("accion", "completar")  # "completar" | "deshacer"
    hoy      = _d.today().isoformat()

    if not clave:
        return redirect(url_for("formacion.formacion_alarmas"))

    conn = get_form_conn()
    if accion == "deshacer":
        conn.execute(
            "DELETE FROM alarmas_completadas WHERE tutor_id=? AND clave=? AND fecha_dia=?",
            (tutor_id, clave, hoy)
        )
    else:
        conn.execute(
            "INSERT IGNORE INTO alarmas_completadas (tutor_id, clave, fecha_dia) VALUES (?,?,?)",
            (tutor_id, clave, hoy)
        )
    conn.commit()
    conn.close()
    return redirect(url_for("formacion.formacion_alarmas"))


@formacion_bp.route("/formacion/alarmas/badge")
@login_required
def alarmas_badge():
    """Endpoint JSON para actualizar el badge sin recargar la página."""
    tutor_id   = session["user_id"]
    alarmas    = _generar_alarmas(tutor_id)
    completadas = _get_completadas_hoy(tutor_id)
    pendientes = sum(1 for a in alarmas if a["clave"] not in completadas)
    return jsonify({"pendientes": pendientes})


# ── Ruta: datos de calendario (FullCalendar) ───────────────────────────────────
@formacion_bp.route("/formacion/calendar-data")
@login_required
def formacion_calendar_data():
    """Devuelve eventos {tipo,fecha,curso,alumno} y notas para el calendario custom."""
    tutor_id = session["user_id"]
    conn = get_form_conn()
    alumnos = [dict(a) for a in conn.execute(
        "SELECT id, nombre, curso, fecha_inicio, fecha_fin "
        "FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0)",
        (tutor_id,)
    ).fetchall()]
    conn.close()

    # Deduplicar: una entrada por (tipo, fecha, curso) — no una por alumno
    vistos  = set()
    eventos = []
    cursos_sin_fecha = set()  # cursos que existen pero sin fechas asignadas

    for a in alumnos:
        curso = a.get("curso") or "Sin curso"
        tiene_fecha = False
        if a.get("fecha_inicio"):
            key = ("inicio", str(a["fecha_inicio"])[:10], curso)
            if key not in vistos:
                vistos.add(key)
                eventos.append({"tipo": "inicio", "fecha": key[1], "curso": curso})
            tiene_fecha = True
        if a.get("fecha_fin"):
            key = ("fin", str(a["fecha_fin"])[:10], curso)
            if key not in vistos:
                vistos.add(key)
                eventos.append({"tipo": "fin", "fecha": key[1], "curso": curso})
            tiene_fecha = True
        if not tiene_fecha:
            cursos_sin_fecha.add(curso)

    # Cursos sin fechas: añadir como evento tipo "sin_fecha" con fecha de hoy
    # para que aparezcan en el Gantt y calendario aunque no tengan fechas asignadas
    from datetime import date as _date
    hoy_str = _date.today().isoformat()
    cursos_con_evento = {e["curso"] for e in eventos}
    for curso in cursos_sin_fecha:
        if curso not in cursos_con_evento:
            eventos.append({"tipo": "sin_fecha", "fecha": hoy_str, "curso": curso})

    # Notas del calendario
    conn2 = get_form_conn()
    conn2.execute("""
        CREATE TABLE IF NOT EXISTS notas_calendario (
            id         INT PRIMARY KEY AUTO_INCREMENT,
            tutor_id   INT         NOT NULL,
            fecha      DATE        NOT NULL,
            nota       TEXT,
            color      VARCHAR(20) DEFAULT 'amber',
            created_at TIMESTAMP   DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uq_tutor_fecha_nota (tutor_id, fecha, nota(100))
        )
    """)
    rows = conn2.execute(
        "SELECT id, fecha, nota, color FROM notas_calendario WHERE tutor_id=? ORDER BY id ASC",
        (tutor_id,)
    ).fetchall()
    conn2.commit()
    conn2.close()
    notas = [{"id": r["id"], "fecha": str(r["fecha"]), "nota": r["nota"], "color": r["color"]} for r in rows]

    return jsonify({"eventos": eventos, "notas": notas})


# ── Rutas: notas de calendario ─────────────────────────────────────────────────
@formacion_bp.route("/formacion/calendar/nota/guardar", methods=["POST"])
@login_required
def guardar_nota_calendario():
    """Guarda o actualiza una nota. El template envía: {fecha, nota, color, nota_id?}."""
    tutor_id = session["user_id"]
    data     = request.get_json(silent=True) or {}
    fecha    = data.get("fecha", "").strip()
    nota     = data.get("nota", "").strip()
    color    = data.get("color", "amber").strip()
    nota_id  = data.get("nota_id")  # None si es nueva

    if not fecha or not nota:
        return jsonify({"ok": False, "error": "Fecha y nota requeridas"}), 400

    conn = get_form_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS notas_calendario (
            id         INT PRIMARY KEY AUTO_INCREMENT,
            tutor_id   INT         NOT NULL,
            fecha      DATE        NOT NULL,
            nota       TEXT,
            color      VARCHAR(20) DEFAULT 'amber',
            created_at TIMESTAMP   DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uq_tutor_fecha_nota (tutor_id, fecha, nota(100))
        )
    """)
    if nota_id:
        conn.execute(
            "UPDATE notas_calendario SET nota=?, color=? WHERE id=? AND tutor_id=?",
            (nota, color, nota_id, tutor_id)
        )
        new_id = nota_id
    else:
        cur = conn.execute(
            "INSERT INTO notas_calendario (tutor_id, fecha, nota, color) VALUES (?,?,?,?)",
            (tutor_id, fecha, nota, color)
        )
        new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@formacion_bp.route("/formacion/calendario/nota/borrar/<int:nota_id>", methods=["POST"])
@login_required
def borrar_nota_calendario(nota_id):
    """Elimina una nota por su id."""
    tutor_id = session["user_id"]
    conn = get_form_conn()
    conn.execute(
        "DELETE FROM notas_calendario WHERE id=? AND tutor_id=?",
        (nota_id, tutor_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ── Ruta: dashboard de formación ───────────────────────────────────────────────
@formacion_bp.route("/formacion/dashboard")
@login_required
def formacion_dashboard():
    tutor_id = session.get("user_id")
    conn     = get_form_conn()

    # Convertir a dicts para que tojson pueda serializarlos en el template
    alumnos = [dict(a) for a in conn.execute(
        "SELECT * FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0) ORDER BY progreso DESC", (tutor_id,)
    ).fetchall()]

    # Adjuntar historial de progreso a cada alumno
    for a in alumnos:
        hist = conn.execute(
            "SELECT fecha_import, progreso, examenes, delta_progreso, avanzo "
            "FROM progreso_historial WHERE alumno_id=? ORDER BY fecha_import ASC",
            (a["id"],)
        ).fetchall()
        a["historial"] = [dict(h) for h in hist]
    conn.close()

    # Normalizar campos que pueden ser None
    for a in alumnos:
        a["progreso"]  = float(a.get("progreso") or 0)
        a["examenes"]  = _fmt_examenes(a.get("examenes"))
        _ep = _parse_examenes(a["examenes"])
        a["ex_pending"] = _ep[2] > 0 and (_ep[0] / _ep[2]) < 0.75  # ratio realizados/totales < 75%
        a["supera_75"] = int(a.get("supera_75") or 0)
        a["curso"]     = a.get("curso") or ""

    # Cursos únicos ordenados
    cursos = sorted(set(a["curso"] for a in alumnos if a["curso"]))

    total          = len(alumnos)
    superan_75     = sum(1 for a in alumnos if a["supera_75"] == 1)
    no_superan     = total - superan_75
    pct_exito      = round(superan_75 / total * 100, 1) if total else 0
    avg_progreso   = round(sum(a["progreso"] for a in alumnos) / total, 1) if total else 0
    total_examenes = sum(_parse_examenes(a["examenes"])[0] for a in alumnos)

    # Snapshots históricos (también como dicts)
    conn2     = get_form_conn()
    snapshots = [dict(s) for s in conn2.execute(
        "SELECT * FROM historial_snapshots WHERE tutor_id=? ORDER BY fecha ASC", (tutor_id,)
    ).fetchall()]
    conn2.close()

    snap_labels = [s["label"]        for s in snapshots]
    snap_pct    = [s["pct_exito"]    for s in snapshots]
    snap_avg    = [s["avg_progreso"] for s in snapshots]
    snap_total  = [s["total"]        for s in snapshots]

    # Historial automático
    conn3    = get_form_conn()
    historial_auto = [dict(h) for h in conn3.execute(
        "SELECT * FROM historial_automatico WHERE tutor_id=? ORDER BY id DESC LIMIT 50",
        (tutor_id,)
    ).fetchall()]
    conn3.close()

    # Datos para gráfico de evolución automática
    hist_labels   = [h["fecha"]         for h in reversed(historial_auto)]
    hist_alumnos  = [h["total_alumnos"] for h in reversed(historial_auto)]
    hist_cursos   = [h["total_cursos"]  for h in reversed(historial_auto)]

    return render_template(
        "formacion_dashboard.html",
        alumnos=alumnos,
        cursos=cursos,
        total=total,
        superan_75=superan_75,
        no_superan=no_superan,
        pct_exito=pct_exito,
        avg_progreso=avg_progreso,
        total_examenes=total_examenes,
        snapshots=snapshots,
        snap_labels=snap_labels,
        snap_pct=snap_pct,
        snap_avg=snap_avg,
        snap_total=snap_total,
        historial_auto=historial_auto,
        hist_labels=hist_labels,
        hist_alumnos=hist_alumnos,
        hist_cursos=hist_cursos,
    )


# ── Ruta: descargar modelo de importación ─────────────────────────────────────
@formacion_bp.route("/formacion/descargar_modelo")
@login_required
def descargar_modelo():
    import os
    from flask import send_file, current_app
    # Buscar el modelo en varias ubicaciones posibles
    posibles = [
        os.path.join(current_app.root_path, "static", "MODELO_IMPORTAR.xlsx"),
        os.path.join(current_app.root_path, "MODELO_IMPORTAR.xlsx"),
        os.path.join(os.path.dirname(__file__), "MODELO_IMPORTAR.xlsx"),
        os.path.join(os.path.dirname(__file__), "static", "MODELO_IMPORTAR.xlsx"),
    ]
    for ruta in posibles:
        if os.path.exists(ruta):
            return send_file(ruta, as_attachment=True, download_name="MODELO_IMPORTAR.xlsx")
    # Si no se encuentra el archivo físico, devolver 404 con mensaje claro
    from flask import abort
    abort(404, "El archivo MODELO_IMPORTAR.xlsx no se encontró en el servidor. "
               "Colócalo en la carpeta 'static/' junto a la aplicación.")


# ── Ruta: API WhatsApp link ────────────────────────────────────────────────────
@formacion_bp.route("/formacion/whatsapp/<int:alumno_id>")
@login_required
def whatsapp_alumno(alumno_id):
    conn   = get_form_conn()
    alumno = conn.execute("SELECT * FROM alumnos WHERE id=?", (alumno_id,)).fetchone()
    conn.close()

    if not alumno:
        return redirect(url_for("formacion.formacion"))

    telefono = (alumno["telefono"] or "").strip().replace(" ", "").replace("-", "").replace("+", "")
    if not telefono:
        return redirect(url_for("formacion.formacion"))

    # Formatear nombre: solo nombre y apellido en title case
    partes = alumno["nombre"].strip().split()
    nombre_corto = " ".join(p.capitalize() for p in partes[:2]) if len(partes) >= 2 else partes[0].capitalize() if partes else alumno["nombre"]

    progreso = float(alumno["progreso"] or 0)
    curso    = alumno["curso"] or ""

    # Formatear fechas como "día de mes de año"
    meses = ["enero","febrero","marzo","abril","mayo","junio",
             "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    def fmt_fecha_wa(val):
        if not val: return None
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(str(val)[:10], "%Y-%m-%d")
            return f"{d.day} de {meses[d.month-1]} de {d.year}"
        except Exception:
            return str(val)

    inicio_fmt = fmt_fecha_wa(alumno["fecha_inicio"])
    fin_fmt    = fmt_fecha_wa(alumno["fecha_fin"])

    # Exámenes R/S/T
    ex_raw  = _fmt_examenes(alumno.get("examenes"))
    ex_p    = _parse_examenes(ex_raw)
    ex_r, ex_s, ex_t = ex_p

    supera = progreso >= 75

    if supera:
        intro    = f"Quería escribirte para compartirte los resultados de tu formación y felicitarte por tu esfuerzo."
        cierre   = f"Estás haciendo un trabajo excelente. ¡Sigue así hasta completarlo al 100%! 💪"
    else:
        intro    = f"Me pongo en contacto contigo para hacerte un seguimiento de tu avance en el curso."
        cierre   = f"Recuerda que el objetivo es alcanzar el 75% de progreso. Si necesitas ayuda, aquí estoy. ¡Tú puedes! 🙌"

    periodo = ""
    if inicio_fmt and fin_fmt:
        periodo = f"📆 Periodo: del {inicio_fmt} al {fin_fmt}"
    elif fin_fmt:
        periodo = f"📆 Fecha límite: {fin_fmt}"
    elif inicio_fmt:
        periodo = f"📆 Inicio: {inicio_fmt}"

    ex_linea = ""
    if ex_t > 0:
        ex_linea = f"📝 Exámenes: {ex_r} realizados · {ex_s} superados · {ex_t} totales"
    elif ex_r > 0:
        ex_linea = f"📝 Exámenes realizados: {ex_r}"

    lineas = [
        f"Hola {nombre_corto},",
        "",
        intro,
        "",
        f"📚 Curso: {curso}" if curso else None,
        periodo if periodo else None,
        f"📊 Progreso actual: *{progreso:.0f}%* {'✅' if supera else '⚠️'}",
        ex_linea if ex_linea else None,
        "",
        cierre,
        "",
        "Un saludo,",
        "Ricardo"
    ]

    mensaje = "\n".join(l for l in lineas if l is not None)

    import urllib.parse
    url = f"https://wa.me/{telefono}?text={urllib.parse.quote(mensaje)}"
    from flask import redirect as redir
    return redir(url)


# ── Ruta: consolidado de alumnos ───────────────────────────────────────────────
@formacion_bp.route("/formacion/consolidado")
@login_required
def ver_consolidado():
    df = consolidar_desde_db(FORM_DB, tutor_id=session["user_id"])
    return df.to_json(orient="records", force_ascii=False)


# ══════════════════════════════════════════════════════════════════════════════
# RUTAS IA — Google Gemini
# ══════════════════════════════════════════════════════════════════════════════

# ── Ruta: sección IA (chatbot + análisis de curso) ────────────────────────────
# ── Ruta: proxy IA — generar mensaje Moodle via Groq ─────────────────────────
@formacion_bp.route("/formacion/ia/generar_mensaje_moodle", methods=["POST"])
@login_required
def ia_generar_mensaje_moodle():
    import os, requests as _req
    from dotenv import load_dotenv as _ldenv
    _ldenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
    datos   = request.get_json(silent=True) or {}
    api_key = os.getenv("GROQ_API_KEY", "")
    if not api_key:
        return jsonify({"error": "GROQ_API_KEY no configurada en .env"}), 500

    nombre   = datos.get("nombre", "")
    curso    = datos.get("curso", "")
    progreso = datos.get("progreso", 0)
    examenes = datos.get("examenes", "0/0/0")
    supera   = datos.get("supera", False)
    inicio   = datos.get("inicio", "")
    fin      = datos.get("fin", "")
    tipo     = datos.get("tipo", "seguimiento")

    meses = ["enero","febrero","marzo","abril","mayo","junio",
             "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    def fmt_f(val):
        if not val: return ""
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(str(val)[:10], "%Y-%m-%d")
            return f"{d.day} de {meses[d.month-1]} de {d.year}"
        except Exception:
            return str(val)

    inicio_fmt = fmt_f(inicio)
    fin_fmt    = fmt_f(fin)

    tipos_desc = {
        "seguimiento":  "seguimiento amable del progreso",
        "felicitacion": "felicitación por superar el 75%",
        "advertencia":  "aviso urgente por progreso bajo",
        "recordatorio": "recordatorio de curso pendiente",
        "libre":        "mensaje libre personalizado",
    }
    tipo_desc = tipos_desc.get(tipo, "seguimiento")

    prompt = f"""Escribe un mensaje de Moodle en español para un alumno de formación online.
Tipo de mensaje: {tipo_desc}

Datos del alumno:
- Nombre: {nombre} (usa solo nombre y primer apellido)
- Curso: {curso}
- Progreso: {progreso}%
- Exámenes: {examenes} (formato realizados/superados/totales)
- Supera el 75%: {'Sí' if supera else 'No'}
{f'- Fecha inicio: {inicio_fmt}' if inicio_fmt else ''}
{f'- Fecha fin: {fin_fmt}' if fin_fmt else ''}

Instrucciones:
- Tono cercano, profesional y positivo
- Menciona fechas en formato "día de mes de año"
- Usa los datos reales del alumno
- Mensaje conciso (máximo 150 palabras)
- Devuelve SOLO un JSON con dos campos: "asunto" y "mensaje"
- Sin markdown, sin explicaciones, solo el JSON
"""

    try:
        resp = _req.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "llama-3.3-70b-versatile",
                "max_tokens": 500,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=20
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"].strip()
        # Limpiar posibles backticks
        content = content.replace("```json", "").replace("```", "").strip()
        import json as _json
        result = _json.loads(content)
        return jsonify({"asunto": result.get("asunto",""), "mensaje": result.get("mensaje","")})
    except _req.exceptions.ConnectionError as e:
        return jsonify({"error": f"Sin conexión a internet o Groq caído: {str(e)}"}), 500
    except _req.exceptions.Timeout:
        return jsonify({"error": "Timeout: Groq tardó demasiado"}), 500
    except _req.exceptions.HTTPError as e:
        return jsonify({"error": f"Error HTTP {e.response.status_code}: {e.response.text[:200]}"}), 500
    except _json.JSONDecodeError as e:
        return jsonify({"error": f"IA devolvió respuesta no válida: {content[:200]}"}), 500
    except Exception as e:
        import traceback
        return jsonify({"error": f"{type(e).__name__}: {str(e)}", "detalle": traceback.format_exc()[-500:]}), 500




@formacion_bp.route("/formacion/alumno/observaciones/<int:alumno_id>", methods=["POST"])
@login_required
def add_observacion(alumno_id):
    tutor_id = session["user_id"]
    texto = (request.get_json(silent=True) or {}).get("texto", "").strip()
    if not texto:
        return jsonify({"error": "Texto vacío"}), 400
    conn = get_form_conn()
    conn.execute(
        "INSERT INTO observaciones_alumno (alumno_id, tutor_id, texto) VALUES (%s, %s, %s)",
        (alumno_id, tutor_id, texto)
    )
    conn.commit()
    # Devolver la observación recién creada
    row = conn.execute(
        "SELECT id, texto, created_at FROM observaciones_alumno "
        "WHERE alumno_id=%s AND tutor_id=%s ORDER BY id DESC LIMIT 1",
        (alumno_id, tutor_id)
    ).fetchone()
    conn.close()
    return jsonify({
        "id":    row["id"],
        "texto": row["texto"],
        "fecha": str(row["created_at"])[:16].replace("T", " ") if row["created_at"] else ""
    })


@formacion_bp.route("/formacion/alumno/observaciones/borrar/<int:obs_id>", methods=["POST"])
@login_required
def del_observacion(obs_id):
    tutor_id = session["user_id"]
    conn = get_form_conn()
    conn.execute(
        "DELETE FROM observaciones_alumno WHERE id=%s AND tutor_id=%s",
        (obs_id, tutor_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ── Observaciones de alumno ────────────────────────────────────────────────────
@formacion_bp.route("/formacion/alumno/observaciones/<int:alumno_id>")
@login_required
def get_observaciones(alumno_id):
    tutor_id = session["user_id"]
    conn = get_form_conn()
    rows = conn.execute(
        "SELECT id, texto, created_at FROM observaciones_alumno "
        "WHERE alumno_id=%s AND tutor_id=%s ORDER BY created_at DESC",
        (alumno_id, tutor_id)
    ).fetchall()
    conn.close()
    return jsonify([{
        "id":    r["id"],
        "texto": r["texto"],
        "fecha": str(r["created_at"])[:16] if r["created_at"] else ""
    } for r in rows])


@formacion_bp.route("/formacion/ia")
@login_required
def formacion_ia():
    tutor_id = session["user_id"]
    conn     = get_form_conn()
    cursos   = [r[0] for r in conn.execute(
        "SELECT DISTINCT curso FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0) AND curso IS NOT NULL ORDER BY curso",
        (tutor_id,)
    ).fetchall()]
    conn.close()
    return render_template("formacion_ia.html", cursos=cursos)


# ── Ruta: chatbot del tutor (JSON) ────────────────────────────────────────────
@formacion_bp.route("/formacion/ia/chat", methods=["POST"])
@login_required
def ia_chat():
    tutor_id = session["user_id"]
    datos    = request.get_json(silent=True) or {}
    mensaje  = datos.get("mensaje", "").strip()
    if not mensaje:
        return jsonify({"error": "Mensaje vacío"}), 400
    respuesta = chatbot_tutor(tutor_id, mensaje)
    return jsonify({"respuesta": respuesta})


# ── Ruta: limpiar chat ────────────────────────────────────────────────────────
@formacion_bp.route("/formacion/ia/chat/limpiar", methods=["POST"])
@login_required
def ia_chat_limpiar():
    limpiar_chat(session["user_id"])
    return jsonify({"ok": True})


# ── Ruta: analizar alumno individual (JSON) ───────────────────────────────────
@formacion_bp.route("/formacion/ia/analizar/<int:alumno_id>")
@login_required
def ia_analizar_alumno(alumno_id):
    resultado = analizar_alumno(alumno_id)
    return jsonify(resultado)


# ── Ruta: generar mensaje WhatsApp con IA (JSON) ──────────────────────────────
@formacion_bp.route("/formacion/ia/mensaje_wa/<int:alumno_id>")
@login_required
def ia_mensaje_wa(alumno_id):
    contexto = request.args.get("contexto", "recordatorio")
    mensaje  = generar_mensaje_wa(alumno_id, contexto)
    return jsonify({"mensaje": mensaje})


# ── Ruta: predicción de riesgo de un curso (JSON) ─────────────────────────────
@formacion_bp.route("/formacion/ia/curso", methods=["POST"])
@login_required
def ia_analizar_curso():
    tutor_id = session["user_id"]
    datos    = request.get_json(silent=True) or {}
    curso    = datos.get("curso", "").strip()
    if not curso:
        return jsonify({"error": "Curso no especificado"}), 400
    resultado = predecir_riesgo_curso(tutor_id, curso)
    return jsonify(resultado)



# ── Ruta: importar teléfonos desde Excel ─────────────────────────────────────
@formacion_bp.route("/formacion/importar_telefonos", methods=["POST"])
@login_required
def importar_telefonos():
    tutor_id = session["user_id"]
    arch     = request.files.get("excel_tel")
    if not arch or arch.filename == "":
        return jsonify({"error": "No se recibió ningún archivo."}), 400
    if not arch.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "El archivo debe ser .xlsx o .xls"}), 400
    arch.seek(0)
    datos = arch.read()
    if len(datos) < 100:
        return jsonify({"error": "El archivo parece estar vacío o dañado."}), 400
    resultado = importar_telefonos_excel(datos, tutor_id, FORM_DB)
    return jsonify(resultado)


# ── Ruta: IA — resumen semanal ────────────────────────────────────────────────
@formacion_bp.route("/formacion/ia/resumen_semanal")
@login_required
def ia_resumen_semanal():
    texto = resumen_semanal(session["user_id"])
    return jsonify({"resumen": texto})


# ── Ruta: IA — ranking de riesgo ──────────────────────────────────────────────
@formacion_bp.route("/formacion/ia/ranking")
@login_required
def ia_ranking():
    data = ranking_riesgo(session["user_id"])
    return jsonify(data)


# ── Ruta: IA — mensajes WA masivos ────────────────────────────────────────────
@formacion_bp.route("/formacion/ia/mensajes_masivos", methods=["POST"])
@login_required
def ia_mensajes_masivos():
    datos  = request.get_json(silent=True) or {}
    filtro = datos.get("filtro", "en_riesgo")
    data   = mensajes_wa_masivos(session["user_id"], filtro)
    return jsonify(data)


# ── Ruta: IA — comparativa de cursos ─────────────────────────────────────────
@formacion_bp.route("/formacion/ia/comparativa")
@login_required
def ia_comparativa():
    data = comparativa_cursos(session["user_id"])
    return jsonify(data)


# ── Ruta: IA — sugerencias del día ────────────────────────────────────────────
@formacion_bp.route("/formacion/ia/sugerencias")
@login_required
def ia_sugerencias():
    data = sugerencias_hoy(session["user_id"])
    return jsonify(data)


# ── Ruta: exportar alumnos a Excel ────────────────────────────────────────────

def _fmt_fecha(val):
    """Convierte YYYY-MM-DD a DD/MM/YYYY, devuelve '—' si vacío."""
    if not val:
        return "—"
    try:
        from datetime import datetime as _dtt
        return _dtt.strptime(str(val)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return str(val)


# ── Ruta: exportar Excel por curso seleccionado ───────────────────────────────
@formacion_bp.route("/formacion/exportar_curso_excel")
@login_required
def exportar_curso_excel():
    import io as _io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    tutor_id = session["user_id"]
    curso    = request.args.get("curso", "").strip()

    conn = get_form_conn()
    query = "SELECT * FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0)"
    params = [tutor_id]
    if curso:
        query += " AND curso=?"
        params.append(curso)
    query += " ORDER BY progreso DESC, nombre ASC"
    alumnos = [dict(a) for a in conn.execute(query, params).fetchall()]
    # Cargar observaciones de todos los alumnos del tutor
    obs_rows = conn.execute(
        "SELECT alumno_id, texto, created_at FROM observaciones_alumno WHERE tutor_id=? ORDER BY created_at ASC",
        (tutor_id,)
    ).fetchall()
    obs_map = {}
    for ob in obs_rows:
        aid = ob["alumno_id"]
        fecha = str(ob["created_at"] or "")[:10]
        linea = f"{fecha}: {ob['texto']}" if fecha else ob["texto"]
        obs_map.setdefault(aid, []).append(linea)
    conn.close()

    C_DARK    = "1E3A5F"
    C_GREEN   = "2D9D78"
    C_AMBER   = "D4A017"
    C_RED_TXT = "C0392B"
    C_RED_BG  = "FDECEA"
    C_ALT     = "F0F4F8"
    C_WHITE   = "FFFFFF"
    C_BORDER  = "CBD5E1"

    def thin():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr(cell, text, bg=None):
        cell.value     = text
        cell.font      = Font(bold=True, color=C_WHITE, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg or C_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin()

    wb = Workbook()
    ws = wb.active
    titulo_curso = curso if curso else "Todos los cursos"
    ws.title = titulo_curso[:28] if len(titulo_curso) > 28 else titulo_curso

    # ── Título ──
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = f"{titulo_curso}  ·  {_dt.now().strftime('%d/%m/%Y')}  ·  {len(alumnos)} alumnos"
    c.font      = Font(bold=True, size=13, color=C_DARK, name="Arial")
    c.fill      = PatternFill("solid", fgColor="E8F0FA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 5

    # ── KPIs en fila 2 ──
    contables  = [a for a in alumnos if not a.get("no_llamar")]
    n_cont     = len(contables)
    superan    = sum(1 for a in contables if a.get("supera_75"))
    pct_exito  = round(superan / n_cont * 100, 1) if n_cont else 0
    avg_prog   = round(sum(float(a.get("progreso") or 0) for a in contables) / n_cont, 1) if n_cont else 0
    no_llamar_n = len(alumnos) - n_cont

    kpis = [
        ("Total alumnos", len(alumnos)),
        ("Superan 75%",   superan),
        ("Tasa éxito",    f"{pct_exito}%"),
        ("Prog. medio",   f"{avg_prog}%"),
        ("No Participa",     no_llamar_n),
    ]
    for c_i, (lbl, val) in enumerate(kpis, 1):
        cl = ws.cell(3, c_i, lbl)
        cl.font      = Font(bold=True, size=8, color="64748B", name="Arial")
        cl.fill      = PatternFill("solid", fgColor="E8F0FA")
        cl.alignment = Alignment(horizontal="center")
        cv = ws.cell(4, c_i, val)
        cv.font      = Font(bold=True, size=13, color=C_DARK, name="Arial")
        cv.fill      = PatternFill("solid", fgColor="E8F0FA")
        cv.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 5

    # ── Cabeceras ──
    COLS = ["#", "Nombre", "Progreso (%)", "Ex. Realizados", "Ex. Superados", "Ex. Totales", "Fecha Inicio", "Fecha Fin", "Supera 75%", "Teléfono", "Estado", "Observaciones"]
    ws.row_dimensions[6].height = 30
    for c_i, h in enumerate(COLS, 1):
        hdr(ws.cell(6, c_i), h, bg="4A6FA5" if h == "Observaciones" else None)

    # ── Datos ──
    for r_i, a in enumerate(alumnos, 7):
        es_no_llamar = bool(a.get("no_llamar"))
        p = float(a.get("progreso") or 0)
        obs_texto = "\n".join(obs_map.get(a.get("id"), []))
        ws.row_dimensions[r_i].height = max(18, 15 * len(obs_map.get(a.get("id"), []))) if obs_texto else 18

        bg_row = C_RED_BG if es_no_llamar else (C_ALT if r_i % 2 == 0 else C_WHITE)
        rf     = PatternFill("solid", fgColor=bg_row)

        prog_color = C_GREEN if p >= 75 else (C_AMBER if p >= 50 else C_RED_TXT)

        def dc(col, val, fmt=None, bold=False, center=False, color=None):
            cell = ws.cell(r_i, col, val)
            cell.fill      = rf
            cell.font      = Font(name="Arial", size=9, bold=bold or es_no_llamar,
                                  color=C_RED_TXT if es_no_llamar else (color or "1E293B"))
            cell.border    = thin()
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
            if fmt: cell.number_format = fmt

        nombre_val = a.get("nombre", "") + (" 🔕 NO Participa" if es_no_llamar else "")
        dc(1, r_i - 6,                                 center=True, color="64748B")
        dc(2, nombre_val,                              bold=True)
        dc(3, p,                                       fmt='0.0"%"', center=True, bold=True, color=prog_color)
        _ep = _parse_examenes(a.get("examenes"))
        _ex_color = "7C3AED" if (_ep[2] > 0 and (_ep[0] / _ep[2]) < 0.75) else None
        dc(4, _ep[0], center=True, color=_ex_color)
        dc(5, _ep[1], center=True, color=_ex_color)
        dc(6, _ep[2], center=True, color=_ex_color)
        dc(7, _fmt_fecha(a.get("fecha_inicio")),        center=True)
        dc(8, _fmt_fecha(a.get("fecha_fin")),           center=True)
        dc(9, "✔ Sí" if a.get("supera_75") else "✖ No", center=True, bold=True,
           color=C_GREEN if a.get("supera_75") else C_AMBER)
        dc(10, a.get("telefono", "—") or "—")
        estado = "✅ Supera 75%" if a.get("supera_75") else "⚠ Bajo 75%"
        dc(11, estado, center=True, bold=True,
           color=C_GREEN if a.get("supera_75") else C_AMBER)
        # Observaciones
        obs_cell = ws.cell(r_i, 12, obs_texto or "—")
        obs_cell.fill      = rf
        obs_cell.font      = Font(name="Arial", size=8, italic=bool(obs_texto), color="334155")
        obs_cell.border    = thin()
        obs_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    last = 6 + len(alumnos)
    ws.auto_filter.ref = f"A6:{get_column_letter(len(COLS))}{last}"
    ws.freeze_panes    = "A7"
    for i, w in enumerate([4, 28, 14, 10, 10, 10, 14, 14, 12, 18, 16, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Leyenda ──
    ley_row = last + 2
    ws.merge_cells(start_row=ley_row, start_column=1, end_row=ley_row, end_column=4)
    lc = ws.cell(ley_row, 1, "🔕 Fondo rojo = alumno marcado como No Participa (excluido de estadísticas)")
    lc.font      = Font(italic=True, size=8, color=C_RED_TXT, name="Arial")
    lc.fill      = PatternFill("solid", fgColor=C_RED_BG)
    lc.alignment = Alignment(horizontal="left")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = titulo_curso.replace("/","_").replace("\\","_").replace(" ","_")[:40]
    fname = f"formacion_{safe}_{_dt.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@formacion_bp.route("/formacion/exportar_excel")
@login_required
def exportar_excel():
    import io as _io
    from datetime import datetime as _dt, date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from flask import send_file

    tutor_id = session["user_id"]
    conn     = get_form_conn()
    alumnos  = [dict(a) for a in conn.execute(
        "SELECT * FROM alumnos WHERE tutor_id=? AND (archivado IS NULL OR archivado=0) ORDER BY curso, nombre", (tutor_id,)
    ).fetchall()]
    # Cargar observaciones
    obs_rows2 = conn.execute(
        "SELECT alumno_id, texto, created_at FROM observaciones_alumno WHERE tutor_id=? ORDER BY created_at ASC",
        (tutor_id,)
    ).fetchall()
    obs_map2 = {}
    for ob in obs_rows2:
        aid = ob["alumno_id"]
        fecha = str(ob["created_at"] or "")[:10]
        linea = f"{fecha}: {ob['texto']}" if fecha else ob["texto"]
        obs_map2.setdefault(aid, []).append(linea)
    conn.close()

    # ── Colores ──
    C_DARK   = "1E3A5F"
    C_GREEN  = "2D9D78"
    C_AMBER  = "D4A017"
    C_RED    = "C0392B"
    C_ALT    = "F0F4F8"
    C_WHITE  = "FFFFFF"
    C_BORDER = "CBD5E1"

    def thin():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr(cell, text, bg=None):
        cell.value     = text
        cell.font      = Font(bold=True, color=C_WHITE, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg or C_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin()

    wb = Workbook()

    # ════════════════════════════════════════
    #  HOJA 1 — Todos los alumnos
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = "Alumnos"

    # Título
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = f"Informe de Formación — {_dt.now().strftime('%d/%m/%Y %H:%M')}  ·  {len(alumnos)} alumnos"
    c.font      = Font(bold=True, size=13, color=C_DARK, name="Arial")
    c.fill      = PatternFill("solid", fgColor="E8F0FA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 6

    # Cabeceras
    COLS = ["#","Curso","Nombre","Progreso (%)","Ex. Realizados","Ex. Superados","Ex. Totales","Fecha Inicio",
            "Fecha Fin","Supera 75%","Teléfono","Estado","Importado","Observaciones"]
    ws.row_dimensions[3].height = 30
    for c_i, h in enumerate(COLS, 1):
        hdr(ws.cell(3, c_i), h, bg="4A6FA5" if h == "Observaciones" else None)

    # Datos
    for r, a in enumerate(alumnos, 4):
        obs_texto2 = "\n".join(obs_map2.get(a.get("id"), []))
        ws.row_dimensions[r].height = max(18, 15 * len(obs_map2.get(a.get("id"), []))) if obs_texto2 else 18
        rf = PatternFill("solid", fgColor=C_ALT if r % 2 == 0 else C_WHITE)
        p  = a.get("progreso", 0) or 0

        def dc(col, val, fmt=None, bold=False, center=False, color=None):
            cell = ws.cell(r, col, val)
            cell.fill      = rf
            cell.font      = Font(name="Arial", size=9, bold=bold, color=color or "1E293B")
            cell.border    = thin()
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
            if fmt: cell.number_format = fmt

        prog_color = C_GREEN if p >= 75 else (C_AMBER if p >= 50 else C_RED)
        dc(1,  r-3,                       center=True, color="64748B")
        dc(2,  a.get("curso","—") or "—")
        dc(3,  a.get("nombre",""),         bold=True)
        dc(4,  p,                          fmt='0.0"%"', center=True, bold=True, color=prog_color)
        _ep2 = _parse_examenes(a.get("examenes"))
        _ex2_color = "7C3AED" if (_ep2[2] > 0 and (_ep2[0] / _ep2[2]) < 0.75) else None
        dc(5,  _ep2[0], center=True, color=_ex2_color)
        dc(6,  _ep2[1], center=True, color=_ex2_color)
        dc(7,  _ep2[2], center=True, color=_ex2_color)
        dc(8,  _fmt_fecha(a.get("fecha_inicio")), center=True)
        dc(9,  _fmt_fecha(a.get("fecha_fin")),    center=True)
        dc(10, "✔ Sí" if a.get("supera_75") else "✖ No", center=True, bold=True,
           color=C_GREEN if a.get("supera_75") else C_AMBER)
        dc(11, a.get("telefono","—") or "—")
        dc(12, "✅ Supera 75%" if a.get("supera_75") else "⚠ Bajo 75%",
           center=True, bold=True,
           color=C_GREEN if a.get("supera_75") else C_AMBER)
        dc(13, (a.get("created_at","") or "")[:10], center=True, color="64748B")
        # Observaciones
        obs_cell2 = ws.cell(r, 14, obs_texto2 or "—")
        obs_cell2.fill      = rf
        obs_cell2.font      = Font(name="Arial", size=8, italic=bool(obs_texto2), color="334155")
        obs_cell2.border    = thin()
        obs_cell2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    last_data = 3 + len(alumnos)
    if len(alumnos) > 0:
        ws.conditional_formatting.add(
            f"D4:D{last_data}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="2D9D78")
        )
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{last_data}"
    ws.freeze_panes    = "A4"
    for i, w in enumerate([5,32,22,14,10,10,10,14,14,12,18,16,18,40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════
    #  HOJA 2 — Resumen por curso
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("Resumen por Curso")
    ws2.merge_cells("A1:H1")
    c2 = ws2["A1"]
    c2.value     = "Resumen de rendimiento por curso"
    c2.font      = Font(bold=True, size=13, color=C_DARK, name="Arial")
    c2.fill      = PatternFill("solid", fgColor="E8F0FA")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    ws2.row_dimensions[2].height = 6

    COLS2 = ["Curso","Total Alumnos","Superan 75%","Bajo 75%",
             "Tasa Éxito (%)","Prog. Promedio (%)","Total Exámenes","Prom. Exámenes"]
    ws2.row_dimensions[3].height = 30
    for c_i, h in enumerate(COLS2, 1):
        hdr(ws2.cell(3, c_i), h)

    from collections import defaultdict
    resumen = defaultdict(lambda: {"total":0,"superan":0,"prog":0,"exam":0})
    for a in alumnos:
        k = a.get("curso") or "Sin curso"
        resumen[k]["total"]  += 1
        resumen[k]["superan"] += int(a.get("supera_75") or 0)
        resumen[k]["prog"]   += float(a.get("progreso") or 0)
        resumen[k]["exam"]   += _parse_examenes(a.get("examenes"))[0]

    totales = {"total":0,"superan":0,"prog":0,"exam":0}
    for r, (curso, d) in enumerate(sorted(resumen.items()), 4):
        ws2.row_dimensions[r].height = 18
        rf2 = PatternFill("solid", fgColor=C_ALT if r % 2 == 0 else C_WHITE)
        n   = d["total"]
        sup = d["superan"]
        pct = round(sup/n*100, 1) if n else 0
        avg = round(d["prog"]/n, 1) if n else 0
        avg_e = round(d["exam"]/n, 1) if n else 0
        totales["total"]  += n
        totales["superan"] += sup
        totales["prog"]   += d["prog"]
        totales["exam"]   += d["exam"]

        for c_i, (val, fmt, center, bold, color) in enumerate([
            (curso,   None,       False, True,  "1E293B"),
            (n,       None,       True,  False, "1E293B"),
            (sup,     None,       True,  True,  C_GREEN),
            (n-sup,   None,       True,  False, C_AMBER),
            (pct,     '0.0"%"',  True,  True,  C_GREEN if pct>=75 else (C_AMBER if pct>=50 else C_RED)),
            (avg,     '0.0"%"',  True,  False, "1E293B"),
            (d["exam"],None,      True,  False, "1E293B"),
            (avg_e,   "0.0",     True,  False, "1E293B"),
        ], 1):
            cell = ws2.cell(r, c_i, val)
            cell.fill      = rf2
            cell.font      = Font(name="Arial", size=9, bold=bold, color=color)
            cell.border    = thin()
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
            if fmt: cell.number_format = fmt

    # Fila totales
    tr = 4 + len(resumen)
    tn  = totales["total"]
    tsu = totales["superan"]
    tp  = round(tsu/tn*100, 1) if tn else 0
    ta  = round(totales["prog"]/tn, 1) if tn else 0
    te  = totales["exam"]
    tae = round(te/tn, 1) if tn else 0
    for c_i, (val, fmt) in enumerate([
        ("TOTAL GENERAL",None),(tn,None),(tsu,None),(tn-tsu,None),
        (tp,'0.0"%"'),(ta,'0.0"%"'),(te,None),(tae,"0.0"),
    ], 1):
        cell = ws2.cell(tr, c_i, val)
        cell.font      = Font(bold=True, color=C_WHITE, name="Arial", size=9)
        cell.fill      = PatternFill("solid", fgColor=C_DARK)
        cell.alignment = Alignment(horizontal="center" if c_i>1 else "left", vertical="center")
        cell.border    = thin()
        if fmt: cell.number_format = fmt

    ws2.freeze_panes = "A4"
    for i, w in enumerate([32,14,14,12,16,18,14,16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════
    #  HOJA 3 — Seguimiento pendiente
    # ════════════════════════════════════════
    ws3 = wb.create_sheet("Seguimiento Pendiente")
    ws3.merge_cells("A1:J1")
    c3 = ws3["A1"]
    c3.value     = "⚠ Alumnos que requieren seguimiento — Progreso inferior al 75%"
    c3.font      = Font(bold=True, size=13, color="7B2D00", name="Arial")
    c3.fill      = PatternFill("solid", fgColor="FFF3E0")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26
    ws3.row_dimensions[2].height = 6

    COLS3 = ["#","Curso","Nombre","Progreso (%)","Exámenes","Fecha Inicio",
             "Fecha Fin","Teléfono","Días restantes","Prioridad"]
    ws3.row_dimensions[3].height = 30
    for c_i, h in enumerate(COLS3, 1):
        hdr(ws3.cell(3, c_i), h, bg="8B2500")

    pendientes = sorted([a for a in alumnos if not a.get("supera_75")],
                        key=lambda x: x.get("progreso",0))
    hoy = _date.today()

    for r, a in enumerate(pendientes, 4):
        ws3.row_dimensions[r].height = 18
        rf3 = PatternFill("solid", fgColor=C_ALT if r%2==0 else C_WHITE)
        try:
            ff   = _date.fromisoformat(a.get("fecha_fin",""))
            dias = (ff - hoy).days
        except:
            dias = None

        if dias is None:       prio, pc = "—",           "64748B"
        elif dias < 0:         prio, pc = "🔴 Vencido",  C_RED
        elif dias <= 14:       prio, pc = "🟡 Urgente",  C_AMBER
        else:                  prio, pc = "🟢 En plazo", C_GREEN

        for c_i, (val, fmt, center, bold, color) in enumerate([
            (r-3,                                   None,      True,  False, "64748B"),
            (a.get("curso","—") or "—",             None,      False, False, "1E293B"),
            (a.get("nombre",""),                    None,      False, True,  "1E293B"),
            (a.get("progreso",0),                   '0.0"%"', True,  True,  C_RED),
            (_fmt_examenes(a.get("examenes")),       None,      True,  False, "1E293B"),
            (a.get("fecha_inicio","—") or "—",      None,      True,  False, "1E293B"),
            (a.get("fecha_fin","—")    or "—",      None,      True,  False, "1E293B"),
            (a.get("telefono","—")     or "—",      None,      False, False, "1E293B"),
            (dias if dias is not None else "—",     None,      True,  False, "1E293B"),
            (prio,                                  None,      True,  True,  pc),
        ], 1):
            cell = ws3.cell(r, c_i, val)
            cell.fill      = rf3
            cell.font      = Font(name="Arial", size=9, bold=bold, color=color)
            cell.border    = thin()
            cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
            if fmt: cell.number_format = fmt

    if pendientes:
        ws3.conditional_formatting.add(
            f"D4:D{3+len(pendientes)}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="C0392B")
        )
    ws3.auto_filter.ref = f"A3:J{3+len(pendientes)}" if pendientes else "A3:J3"
    ws3.freeze_panes    = "A4"
    for i, w in enumerate([5,32,22,12,10,14,14,18,14,14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Serializar y enviar ──
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"formacion_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

# ── Ruta: guardar gestión de alumno ──────────────────────────────────────────
@formacion_bp.route("/formacion/alumno_gestion/<int:alumno_id>", methods=["POST"])
@login_required
def alumno_gestion(alumno_id):
    tutor_id = session["user_id"]
    data     = request.get_json(force=True) or {}

    gestionado   = int(data.get("gestionado", 0))
    tipo_gestion = str(data.get("tipo_gestion", "") or "").strip()[:50]
    comentario   = str(data.get("comentario",  "") or "").strip()
    fecha_gestion = datetime.now().strftime("%Y-%m-%d %H:%M")

    conn = get_form_conn()
    conn.execute(
        """UPDATE alumnos
           SET gestionado=?, tipo_gestion=?, comentario=?, fecha_gestion=?
           WHERE id=? AND tutor_id=?""",
        (gestionado, tipo_gestion, comentario, fecha_gestion, alumno_id, tutor_id)
    )
    conn.commit()
    conn.close()

    return jsonify({"ok": True})