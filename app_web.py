from flask import Flask, render_template, request, redirect, session, send_file, url_for
import mysql.connector.errors
from functools import wraps
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from db_mysql import get_tareas_conn, column_exists
from formacion import formacion_bp, inicializar_formacion

app = Flask(__name__)
app.secret_key = "clave_secreta_muy_segura"
app.register_blueprint(formacion_bp)

# ── Perfiles ───────────────────────────────────────────────────────────────────
# 0 = Usuario  |  1 = Admin  |  2 = SuperAdmin

# ── Conexión ───────────────────────────────────────────────────────────────────
def get_connection():
    return get_tareas_conn()

# ── Inicialización ─────────────────────────────────────────────────────────────
def inicializar_todo():
    conn   = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id       INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(150) UNIQUE NOT NULL,
            email    VARCHAR(150) UNIQUE,
            password VARCHAR(255) NOT NULL DEFAULT '',
            es_admin INT DEFAULT 0
        )
    """)
    for col, ddl in [
        ("email",    "VARCHAR(150)"),
        ("es_admin", "INT DEFAULT 0"),
    ]:
        if not column_exists(cursor, 'usuarios', col):
            cursor.execute(f"ALTER TABLE usuarios ADD COLUMN {col} {ddl}")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            descripcion TEXT NOT NULL,
            categoria   VARCHAR(100),
            fecha       VARCHAR(20),
            completada  INT DEFAULT 0,
            codigo      VARCHAR(100),
            usuario_id  INT,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        )
    """)
    for col, ddl in [
        ("usuario_id", "INT"),
        ("prioridad",  "INT DEFAULT 2"),
        ("favorita",   "INT DEFAULT 0"),
        ("notas",      "TEXT"),
        ("usuario",    "VARCHAR(100)"),
    ]:
        if not column_exists(cursor, 'tareas', col):
            cursor.execute(f"ALTER TABLE tareas ADD COLUMN {col} {ddl}")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS subtareas (
            id       INT AUTO_INCREMENT PRIMARY KEY,
            tarea_id INT NOT NULL,
            texto    TEXT NOT NULL,
            hecha    INT DEFAULT 0,
            FOREIGN KEY (tarea_id) REFERENCES tareas(id) ON DELETE CASCADE
        )
    """)

    # Crear admin si no existe
    row = conn.execute(
        "SELECT id FROM usuarios WHERE username='admin'"
    ).fetchone()
    if not row:
        conn.execute(
            "INSERT INTO usuarios (username, email, password, es_admin) VALUES (?,?,?,?)",
            ("admin", "admin@correo.com", generate_password_hash("Admin1234!"), 2)
        )

    conn.commit()
    conn.close()

    # Hashear contraseñas en plano y asegurar es_admin=2 para admin
    # (usa cursor wrapper para que ? → %s funcione correctamente)
    _hashear_passwords()
    print("✅ MySQL tareas inicializada. SuperAdmin listo.")


def _hashear_passwords():
    """Hashea todas las contraseñas en texto plano. Seguro ejecutar múltiples veces."""
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, password FROM usuarios")
    users  = cursor.fetchall()
    for u in users:
        pw = u["password"] or ""
        if pw and not pw.startswith("pbkdf2:") and not pw.startswith("scrypt:"):
            cursor.execute(
                "UPDATE usuarios SET password=? WHERE id=?",
                (generate_password_hash(pw), u["id"])
            )
    cursor.execute("UPDATE usuarios SET es_admin=2 WHERE username=?", ("admin",))
    conn.commit()
    conn.close()

# ── Decoradores ────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Admin (1) y SuperAdmin (2)"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("es_admin", 0) < 1:
            return redirect("/")
        return f(*args, **kwargs)
    return decorated

def superadmin_required(f):
    """Solo SuperAdmin (2)"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("es_admin", 0) != 2:
            return redirect("/")
        return f(*args, **kwargs)
    return decorated

# ── LOGIN / REGISTRO / LOGOUT ──────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        ident    = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        conn     = get_connection()
        usuario  = conn.execute(
            "SELECT id, username, es_admin, password FROM usuarios WHERE username=? OR email=?",
            (ident, ident)
        ).fetchone()
        conn.close()
        if usuario and check_password_hash(dict(usuario)["password"], password):
            session["user_id"]  = usuario["id"]
            session["user"]     = usuario["username"]
            session["es_admin"] = usuario["es_admin"]
            return redirect("/")
        error = "Usuario o contraseña incorrectos."
    return render_template("login.html", error=error)


@app.route("/registro", methods=["GET", "POST"])
def registro():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email    = request.form.get("email",    "").strip()
        password = request.form.get("password", "").strip()
        if not username:
            error = "El nombre de usuario es obligatorio."
        elif len(password) < 6:
            error = "La contraseña debe tener al menos 6 caracteres."
        else:
            conn = get_connection()
            try:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT INTO usuarios (username, email, password, es_admin) VALUES (?,?,?,0)",
                    (username, email, generate_password_hash(password))
                )
                conn.commit()
                conn.close()
                return redirect(url_for("login") + "?registered=1")
            except Exception as e:
                conn.close()
                err_str = str(e).lower()
                if "duplicate" in err_str or "unique" in err_str:
                    error = "El usuario o email ya existe."
                else:
                    error = f"Error al registrar: {e}"
    return render_template("registro.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── RUTAS PRINCIPALES ──────────────────────────────────────────────────────────

@app.route("/accesos_rapidos")
@login_required
def accesos_rapidos():
    return render_template("accesos_rapidos.html")


@app.route("/")
@login_required
def index():
    user_id  = session.get("user_id")
    es_admin = session.get("es_admin", 0)

    filtro_estado = request.args.get("estado", "all").strip()
    filtro_cat    = request.args.get("cat",    "").strip()
    filtro_q      = request.args.get("q",      "").strip()
    filtro_prio   = request.args.get("prio",   "").strip()
    filtro_fav    = request.args.get("fav",    "").strip()
    page     = max(request.args.get("page", 1, type=int), 1)
    per_page = 10

    filtros, params = [], []
    if es_admin < 1:
        filtros.append("usuario_id = ?"); params.append(user_id)
    if filtro_estado == "pending":
        filtros.append("completada = 0")
    elif filtro_estado == "done":
        filtros.append("completada = 1")
    if filtro_cat:
        filtros.append("LOWER(TRIM(COALESCE(categoria,''))) = LOWER(TRIM(?))"); params.append(filtro_cat)
    if filtro_prio in ("1","2","3"):
        filtros.append("prioridad = ?"); params.append(int(filtro_prio))
    if filtro_fav == "1":
        filtros.append("favorita = 1")
    if filtro_q:
        filtros.append("(LOWER(descripcion) LIKE ? OR LOWER(COALESCE(codigo,'')) LIKE ? OR LOWER(COALESCE(categoria,'')) LIKE ?)")
        like = f"%{filtro_q.lower()}%"; params.extend([like, like, like])

    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
    conn  = get_connection()

    total_row    = conn.execute(f"SELECT COUNT(*) AS cnt FROM tareas {where}", params).fetchone()
    total_tareas = total_row["cnt"] if total_row else 0
    total_pages  = max((total_tareas + per_page - 1) // per_page, 1)
    page   = min(page, total_pages)
    offset = (page - 1) * per_page

    tareas = conn.execute(
        f"SELECT * FROM tareas {where} ORDER BY favorita DESC, prioridad ASC, id DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()

    ids = [t["id"] for t in tareas]
    subtareas_map = {}
    if ids:
        ph = ",".join("?" * len(ids))
        for s in conn.execute(f"SELECT * FROM subtareas WHERE tarea_id IN ({ph}) ORDER BY id", ids).fetchall():
            subtareas_map.setdefault(s["tarea_id"], []).append(dict(s))

    if es_admin >= 1:
        cats = conn.execute("SELECT DISTINCT COALESCE(NULLIF(TRIM(categoria),''),'General') AS cat FROM tareas ORDER BY cat").fetchall()
    else:
        cats = conn.execute("SELECT DISTINCT COALESCE(NULLIF(TRIM(categoria),''),'General') AS cat FROM tareas WHERE usuario_id=? ORDER BY cat", (user_id,)).fetchall()
    categorias_lista = [r["cat"] for r in cats]
    conn.close()

    return render_template("index.html", tareas=tareas, page=page, total_pages=total_pages,
                           filtro_estado=filtro_estado, filtro_cat=filtro_cat, filtro_q=filtro_q,
                           filtro_prio=filtro_prio, filtro_fav=filtro_fav,
                           categorias_lista=categorias_lista, subtareas_map=subtareas_map)


@app.route("/agregar", methods=["POST"])
@login_required
def agregar():
    conn = get_connection()
    conn.execute("""
        INSERT INTO tareas (codigo, descripcion, categoria, fecha, completada, usuario_id, prioridad, favorita, notas)
        VALUES (?,?,?,?,0,?,?,0,?)
    """, (request.form.get("codigo"), request.form.get("descripcion"),
          request.form.get("categoria"), request.form.get("fecha"),
          session["user_id"], int(request.form.get("prioridad", 2)),
          request.form.get("notas", "")))
    conn.commit(); conn.close()
    return redirect("/")


@app.route("/admin")
@admin_required
def admin():
    filtro_cat = request.args.get("categoria", "").strip()
    filtro_est = request.args.get("estado", "").strip()
    page       = max(request.args.get("page", 1, type=int), 1)
    per_page   = 10

    filtros, params = [], []
    if filtro_cat:
        filtros.append("LOWER(TRIM(categoria)) = LOWER(TRIM(?))"); params.append(filtro_cat)
    if filtro_est == "Completada":
        filtros.append("completada = 1")
    elif filtro_est == "Pendiente":
        filtros.append("completada = 0")
    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""

    with get_connection() as conn:
        total_row = conn.execute(f"SELECT COUNT(*) AS cnt FROM tareas {where}", params).fetchone()
        total_filtrado = total_row["cnt"] if total_row else 0
        total_pages = max((total_filtrado + per_page - 1) // per_page, 1)
        page   = min(page, total_pages)
        offset = (page - 1) * per_page

        tareas = conn.execute(f"""
            SELECT id, descripcion, categoria, fecha, completada, codigo, usuario_id
            FROM tareas {where} ORDER BY id DESC LIMIT ? OFFSET ?
        """, params + [per_page, offset]).fetchall()

        cats = conn.execute("""
            SELECT DISTINCT categoria FROM tareas
            WHERE categoria IS NOT NULL AND categoria != '' ORDER BY categoria
        """).fetchall()
        categorias_lista = [r["categoria"] for r in cats]

    return render_template("admin.html", tareas=tareas, page=page, total_pages=total_pages,
                           total=total_filtrado, categorias=categorias_lista,
                           filtro_cat=filtro_cat, filtro_est=filtro_est)


# ── GESTIÓN DE USUARIOS ────────────────────────────────────────────────────────

@app.route("/usuarios")
@admin_required
def usuarios():
    conn = get_connection()
    usuarios_lista = conn.execute("""
        SELECT u.id, u.username, u.email, u.es_admin,
               COUNT(t.id)                                      AS total_tareas,
               SUM(CASE WHEN t.completada=1 THEN 1 ELSE 0 END) AS completadas,
               SUM(CASE WHEN t.completada=0 THEN 1 ELSE 0 END) AS pendientes
        FROM usuarios u LEFT JOIN tareas t ON t.usuario_id = u.id
        GROUP BY u.id, u.username, u.email, u.es_admin
        ORDER BY u.es_admin DESC, u.username
    """).fetchall()
    cats = conn.execute("""
        SELECT DISTINCT COALESCE(NULLIF(TRIM(categoria),''),'General') AS cat FROM tareas ORDER BY cat
    """).fetchall()
    categorias = [r["cat"] for r in cats]
    conn.close()
    return render_template("usuarios.html", usuarios=usuarios_lista, categorias=categorias)


@app.route("/usuarios/eliminar/<int:uid>", methods=["POST"])
@admin_required
def usuario_eliminar(uid):
    if uid == session["user_id"]:
        return redirect("/usuarios")
    conn = get_connection()
    conn.execute("DELETE FROM tareas   WHERE usuario_id = ?", (uid,))
    conn.execute("DELETE FROM usuarios WHERE id = ?",         (uid,))
    conn.commit(); conn.close()
    return redirect("/usuarios")


@app.route("/usuarios/asignar_tarea/<int:uid>", methods=["POST"])
@admin_required
def usuario_asignar_tarea(uid):
    descripcion = request.form.get("descripcion", "").strip()
    if not descripcion:
        return redirect("/usuarios")
    conn = get_connection()
    conn.execute("""
        INSERT INTO tareas (codigo, descripcion, categoria, fecha, completada, usuario_id, prioridad, favorita, notas)
        VALUES (?,?,?,?,0,?,?,0,?)
    """, (request.form.get("codigo") or None, descripcion,
          request.form.get("categoria") or None, request.form.get("fecha") or None,
          uid, int(request.form.get("prioridad", 2)), request.form.get("notas") or None))
    conn.commit(); conn.close()
    return redirect("/usuarios")


@app.route("/usuarios/toggle_admin/<int:uid>", methods=["POST"])
@superadmin_required
def usuario_toggle_admin(uid):
    if uid == session["user_id"]:
        return redirect("/usuarios")
    conn = get_connection()
    conn.execute("UPDATE usuarios SET es_admin = 1 - es_admin WHERE id=?", (uid,))
    conn.commit(); conn.close()
    return redirect("/usuarios")


# ── OPERACIONES TAREAS ─────────────────────────────────────────────────────────

@app.route("/completar/<int:id>")
@login_required
def completar(id):
    conn = get_connection()
    if session.get("es_admin", 0) >= 1:
        conn.execute("UPDATE tareas SET completada=1 WHERE id=?", (id,))
    else:
        conn.execute("UPDATE tareas SET completada=1 WHERE id=? AND usuario_id=?", (id, session["user_id"]))
    conn.commit(); conn.close()
    return redirect("/")


@app.route("/eliminar/<int:id>")
@login_required
def eliminar(id):
    conn = get_connection()
    if session.get("es_admin", 0) >= 1:
        conn.execute("DELETE FROM tareas WHERE id=?", (id,))
    else:
        conn.execute("DELETE FROM tareas WHERE id=? AND usuario_id=?", (id, session["user_id"]))
    conn.commit(); conn.close()
    return redirect("/")


@app.route("/favorita/<int:id>")
@login_required
def favorita(id):
    conn = get_connection()
    if session.get("es_admin", 0) >= 1:
        conn.execute("UPDATE tareas SET favorita=1-favorita WHERE id=?", (id,))
    else:
        conn.execute("UPDATE tareas SET favorita=1-favorita WHERE id=? AND usuario_id=?", (id, session["user_id"]))
    conn.commit(); conn.close()
    return redirect(request.referrer or "/")


@app.route("/duplicar/<int:id>")
@login_required
def duplicar(id):
    conn = get_connection()
    t = conn.execute("SELECT * FROM tareas WHERE id=?", (id,)).fetchone()
    if t:
        conn.execute(
            "INSERT INTO tareas (descripcion,categoria,fecha,completada,codigo,usuario_id,prioridad,favorita,notas) VALUES (?,?,?,0,?,?,?,0,?)",
            (t["descripcion"]+" (copia)", t["categoria"], t["fecha"], t["codigo"],
             t["usuario_id"], t["prioridad"] or 2, t["notas"] or "")
        )
        conn.commit()
    conn.close()
    return redirect(request.referrer or "/")


@app.route("/editar/<int:id>", methods=["GET","POST"])
@login_required
def editar(id):
    conn = get_connection()
    if request.method == "POST":
        conn.execute(
            "UPDATE tareas SET codigo=?,descripcion=?,categoria=?,fecha=?,completada=?,prioridad=?,notas=? WHERE id=?",
            (request.form.get("codigo"), request.form.get("descripcion"),
             request.form.get("categoria"), request.form.get("fecha"),
             request.form.get("completada"), int(request.form.get("prioridad", 2)),
             request.form.get("notas",""), id)
        )
        conn.commit(); conn.close()
        return redirect("/")
    tarea = conn.execute("SELECT * FROM tareas WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template("editar.html", tarea=tarea)


@app.route("/subtarea/agregar/<int:tarea_id>", methods=["POST"])
@login_required
def subtarea_agregar(tarea_id):
    texto = request.form.get("texto","").strip()
    if texto:
        conn = get_connection()
        conn.execute("INSERT INTO subtareas (tarea_id,texto) VALUES (?,?)", (tarea_id, texto))
        conn.commit(); conn.close()
    return redirect(request.referrer or "/")


@app.route("/subtarea/toggle/<int:sub_id>")
@login_required
def subtarea_toggle(sub_id):
    conn = get_connection()
    conn.execute("UPDATE subtareas SET hecha=1-hecha WHERE id=?", (sub_id,))
    conn.commit(); conn.close()
    return redirect(request.referrer or "/")


@app.route("/subtarea/eliminar/<int:sub_id>")
@login_required
def subtarea_eliminar(sub_id):
    conn = get_connection()
    conn.execute("DELETE FROM subtareas WHERE id=?", (sub_id,))
    conn.commit(); conn.close()
    return redirect(request.referrer or "/")


# ── DASHBOARD ──────────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    user_id  = session.get("user_id")
    es_admin = session.get("es_admin", 0)
    conn     = get_connection()

    filtro = "" if es_admin >= 1 else "WHERE usuario_id = ?"
    params = () if es_admin >= 1 else (user_id,)

    stats = conn.execute(f"""
        SELECT COUNT(*) AS total, SUM(CASE WHEN completada=1 THEN 1 ELSE 0 END) AS completadas
        FROM tareas {filtro}
    """, params).fetchone()

    total       = stats["total"]       or 0
    completadas = stats["completadas"] or 0
    pendientes  = total - completadas
    porcentaje  = round((completadas / total) * 100, 1) if total > 0 else 0

    if porcentaje > 70:   nivel, color_nivel = "Alto",  "success"
    elif porcentaje > 40: nivel, color_nivel = "Medio", "warning"
    else:                 nivel, color_nivel = "Bajo",  "danger"

    ultimas_tareas = conn.execute(
        f"SELECT * FROM tareas {filtro} ORDER BY id DESC LIMIT 5", params
    ).fetchall()

    datos_cats = conn.execute(
        f"SELECT categoria, COUNT(*) AS cantidad FROM tareas {filtro} GROUP BY categoria ORDER BY cantidad DESC",
        params
    ).fetchall()

    categorias = [r["categoria"] if r["categoria"] else "General" for r in datos_cats]
    cantidades = [r["cantidad"] for r in datos_cats]

    todas_tareas = [
        dict(r) for r in conn.execute(
            f"""SELECT id, descripcion,
            COALESCE(NULLIF(TRIM(categoria),''),'General') AS categoria,
            prioridad, completada, codigo, fecha, notas
            FROM tareas {filtro}
            ORDER BY completada ASC, id DESC""",
            params
        ).fetchall()
    ]

    conn.close()

    return render_template("dashboard.html", total=total, completadas=completadas,
                           pendientes=pendientes, porcentaje=porcentaje, nivel=nivel,
                           color_nivel=color_nivel, fecha_actual=datetime.now().strftime("%d/%m/%Y"),
                           ultimas_tareas=ultimas_tareas, categorias=categorias, cantidades=cantidades,
                           todas_tareas=todas_tareas)


# ── SUPERADMIN: GRÁFICOS GLOBALES ──────────────────────────────────────────────

@app.route("/superadmin")
@superadmin_required
def superadmin():
    conn  = get_connection()
    stats = conn.execute("""
        SELECT COUNT(*) AS total,
               SUM(CASE WHEN completada=1 THEN 1 ELSE 0 END) AS completadas,
               SUM(CASE WHEN completada=0 THEN 1 ELSE 0 END) AS pendientes
        FROM tareas
    """).fetchone()
    por_usuario = [dict(r) for r in conn.execute("""
        SELECT u.username,
               COUNT(t.id)                                      AS total,
               SUM(CASE WHEN t.completada=1 THEN 1 ELSE 0 END) AS completadas,
               SUM(CASE WHEN t.completada=0 THEN 1 ELSE 0 END) AS pendientes
        FROM usuarios u
        LEFT JOIN tareas t ON t.usuario_id = u.id
        GROUP BY u.id, u.username
        ORDER BY total DESC
    """).fetchall()]
    por_categoria = [dict(r) for r in conn.execute("""
        SELECT COALESCE(NULLIF(TRIM(categoria),''), 'General') AS categoria,
               COUNT(*) AS total
        FROM tareas GROUP BY categoria ORDER BY total DESC
    """).fetchall()]
    por_fecha = [dict(r) for r in conn.execute("""
        SELECT fecha, COUNT(*) AS total FROM tareas
        WHERE completada=1 AND fecha IS NOT NULL AND fecha != ''
        GROUP BY fecha ORDER BY fecha DESC LIMIT 30
    """).fetchall()]
    por_fecha.reverse()

    # ── Detalle de tareas para popups ──────────────────────────────────────────
    filas_detalle = conn.execute("""
        SELECT t.descripcion,
               COALESCE(NULLIF(TRIM(t.categoria),''), 'General') AS categoria,
               t.fecha,
               t.completada,
               t.codigo,
               u.username
        FROM tareas t
        JOIN usuarios u ON t.usuario_id = u.id
        ORDER BY t.completada ASC, t.fecha DESC
    """).fetchall()

    from collections import defaultdict
    _det_usuario   = defaultdict(list)
    _det_categoria = defaultdict(list)
    for r in filas_detalle:
        item = {
            "descripcion": r[0] or "",
            "categoria":   r[1] or "General",
            "fecha":       r[2] or "",
            "completada":  int(r[3] or 0),
            "codigo":      r[4] or "",
            "username":    r[5] or ""
        }
        _det_usuario[r[5]].append(item)
        _det_categoria[r[1]].append(item)

    conn.close()
    total       = stats["total"]       or 0
    completadas = stats["completadas"] or 0
    pendientes  = stats["pendientes"]  or 0
    pct         = round(completadas / total * 100, 1) if total else 0
    return render_template("superadmin.html",
        total=total, completadas=completadas, pendientes=pendientes, pct=pct,
        por_usuario=por_usuario, por_categoria=por_categoria, por_fecha=por_fecha,
        detalle_por_usuario=dict(_det_usuario),
        detalle_por_categoria=dict(_det_categoria),
        fecha_actual=datetime.now().strftime("%d/%m/%Y %H:%M"))


# ── EXPORTAR EXCEL ─────────────────────────────────────────────────────────────

_DARK="1A1410"; _DARK2="241C16"; _ACCENT="E6A15A"; _AMBER="D18B47"
_BLUE="C47A5A"; _PURPLE="B07A63"; _TEAL="8F6B50"
_TXT_W="F5EDE6"; _TXT_M="B7A79A"; _BDR="3A2E25"
_CAT_COLS=[("2A1F18","E6A15A"),("332419","D18B47"),("2E1E17","C47A5A"),
           ("2B1B16","B07A63"),("241A15","8F6B50"),("3A221C","E07A5F"),
           ("2F2018","F2A65A"),("352317","C97B3C")]

def _xb():
    s=Side(style="thin",color=_BDR); return Border(left=s,right=s,top=s,bottom=s)
def _xf(h): return PatternFill(start_color=h,end_color=h,fill_type="solid")
def _xfn(bold=False,color=_TXT_W,size=10,italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name="Calibri")
def _xal(h="center",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def _xhdr(cell,bg=_DARK,fg=_TXT_W,size=10):
    cell.fill=_xf(bg); cell.font=_xfn(bold=True,color=fg,size=size)
    cell.alignment=_xal(); cell.border=_xb()
def _xaw(ws,min_w=8,max_w=50):
    for cc in ws.iter_cols():
        f=cc[0]
        if isinstance(f,MergedCell): continue
        ln=max((len(str(c.value)) if c.value is not None and not isinstance(c,MergedCell) else 0) for c in cc)
        ws.column_dimensions[f.column_letter].width=min(max(ln+3,min_w),max_w)

def _hoja_resumen(wb,todas,hoy_str,admin_usr,username):
    ws=wb.active; ws.title="Resumen General"; ws.sheet_properties.tabColor=_ACCENT; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:H1"); c=ws["A1"]
    c.value=f"  REPORTE DE TAREAS — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill=_xf(_DARK); c.font=_xfn(bold=True,size=14,color=_ACCENT); c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:H2"); c=ws["A2"]
    c.value=f"  Generado por: {username}   ·   {'Vista administrador' if admin_usr else 'Vista personal'}"
    c.fill=_xf(_DARK2); c.font=_xfn(color=_TXT_M,size=9,italic=True); c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[2].height=20
    total=len(todas); comp=sum(1 for t in todas if t["completada"]==1); pend=total-comp
    pct=round(comp/total*100,1) if total else 0; n_cats=len({t["categoria"] or "General" for t in todas})
    kpis=[("TOTAL TAREAS",total,_BLUE,"0D2040"),("COMPLETADAS",comp,_ACCENT,"1A4A24"),
          ("PENDIENTES",pend,_AMBER,"3D2E00"),(f"% COMPLETADO",f"{pct}%",_TEAL,"0A2E2A"),("CATEGORÍAS",n_cats,_PURPLE,"1E1040")]
    for rh,h in [(3,10),(4,22),(5,40),(6,18),(7,10)]: ws.row_dimensions[rh].height=h
    for ci,(label,valor,fg,bg) in enumerate(kpis,start=1):
        for rn,val,fsz in [(4,label,8),(5,valor,22),(6,"",9)]:
            c=ws.cell(row=rn,column=ci,value=val); c.fill=_xf(bg); c.font=_xfn(bold=True,color=fg,size=fsz); c.alignment=_xal(); c.border=_xb()
        ws.column_dimensions[get_column_letter(ci)].width=18
    ws.row_dimensions[8].height=22
    for ci,h in enumerate(["Categoría","Total","Completadas","Pendientes","% Completado","Usuarios"],start=1):
        _xhdr(ws.cell(row=8,column=ci,value=h),bg=_DARK,fg=_ACCENT)
    cat_data=defaultdict(lambda:{"total":0,"completadas":0,"usuarios":set()})
    for t in todas:
        cat=t["categoria"] or "General"; cat_data[cat]["total"]+=1; cat_data[cat]["completadas"]+=(1 if t["completada"]==1 else 0)
        if t.get("username"): cat_data[cat]["usuarios"].add(t["username"])
    rn=9
    for ci2,(cat,d) in enumerate(sorted(cat_data.items())):
        bg_f,fg_c=_CAT_COLS[ci2%len(_CAT_COLS)]; pc2=d["total"]-d["completadas"]; ptc2=round(d["completadas"]/d["total"]*100,1) if d["total"] else 0
        u_str=", ".join(sorted(d["usuarios"])) if d["usuarios"] else "—"
        for cj,val in enumerate([cat,d["total"],d["completadas"],pc2,f"{ptc2}%",u_str],start=1):
            c=ws.cell(row=rn,column=cj,value=val); c.fill=_xf(bg_f if cj==1 else _DARK2)
            c.font=_xfn(bold=(cj==1),color=fg_c if cj==1 else _TXT_W); c.alignment=_xal(h="left" if cj in(1,6) else "center"); c.border=_xb()
        ws.row_dimensions[rn].height=18; rn+=1
    for cj,val in enumerate(["TOTAL",total,comp,pend,f"{pct}%",""],start=1):
        c=ws.cell(row=rn,column=cj,value=val); c.fill=_xf(_DARK); c.font=_xfn(bold=True,color=_ACCENT,size=10)
        c.alignment=_xal(h="left" if cj==1 else "center"); c.border=_xb()
    ws.row_dimensions[rn].height=20
    for i,w in enumerate([28,10,14,12,16,32],start=1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A9"

def _hoja_categoria(wb,cat_nombre,tareas_cat,color_idx):
    bg_fill,fg_col=_CAT_COLS[color_idx%len(_CAT_COLS)]; titulo=cat_nombre[:28]+"..." if len(cat_nombre)>31 else cat_nombre
    ws=wb.create_sheet(title=titulo); ws.sheet_properties.tabColor=fg_col; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:G1"); c=ws["A1"]; c.value=f"  {cat_nombre.upper()}"
    c.fill=_xf(bg_fill); c.font=_xfn(bold=True,size=13,color=fg_col); c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[1].height=32
    tc=len(tareas_cat); cc=sum(1 for t in tareas_cat if t["completada"]==1); pc=tc-cc; ptc=round(cc/tc*100,1) if tc else 0
    ws.merge_cells("A2:G2"); c=ws["A2"]
    c.value=f"  Total: {tc}  ·  Completadas: {cc}  ·  Pendientes: {pc}  ·  Progreso: {ptc}%"
    c.fill=_xf(_DARK2); c.font=_xfn(color=_TXT_M,size=9,italic=True); c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[2].height=18
    for ci,h in enumerate(["ID","Código","Descripción","Fecha","Estado","Usuario","Notas"],start=1):
        _xhdr(ws.cell(row=3,column=ci,value=h),bg=bg_fill,fg=fg_col)
    ws.row_dimensions[3].height=20
    for ri,t in enumerate(tareas_cat,start=4):
        done=t["completada"]==1; est="✔  Completada" if done else "●  Pendiente"; ec=_ACCENT if done else _AMBER; rbg="1A4A24" if done else "3D2E00"
        vals=[t["id"],t["codigo"] or "—",t["descripcion"],t["fecha"] or "—",est,t.get("username") or "—",""]
        for ci,val in enumerate(vals,start=1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=_xf(rbg if ci==5 else _DARK2)
            c.font=_xfn(bold=(ci==5),color=ec if ci==5 else(_TXT_M if ci in(1,4,6) else _TXT_W),size=9 if ci in(1,4,6) else 10)
            c.alignment=_xal(h="center" if ci in(1,4,5) else "left",wrap=(ci==3)); c.border=_xb()
        ws.row_dimensions[ri].height=16
    _xaw(ws); ws.column_dimensions["C"].width=45; ws.column_dimensions["G"].width=20; ws.freeze_panes="A4"; ws.auto_filter.ref=ws.dimensions

def _hoja_hoy(wb,tareas_hoy,hoy_str):
    ws=wb.create_sheet(title="Hoy"); ws.sheet_properties.tabColor=_BLUE; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:G1"); c=ws["A1"]
    c.value=f"  TAREAS DEL DÍA — {datetime.strptime(hoy_str,'%Y-%m-%d').strftime('%d / %m / %Y')}"
    c.fill=_xf("0D2040"); c.font=_xfn(bold=True,size=13,color=_BLUE); c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[1].height=32
    th=len(tareas_hoy); ch=sum(1 for t in tareas_hoy if t["completada"]==1)
    ws.merge_cells("A2:G2"); c=ws["A2"]
    c.value=f"  {th} tarea{'s' if th!=1 else ''} programadas hoy  ·  {ch} completadas  ·  {th-ch} pendientes"
    c.fill=_xf(_DARK2); c.font=_xfn(color=_TXT_M,size=9,italic=True); c.alignment=_xal(h="left"); c.border=_xb(); ws.row_dimensions[2].height=18
    for ci,h in enumerate(["ID","Código","Descripción","Categoría","Estado","Usuario","Notas"],start=1):
        _xhdr(ws.cell(row=3,column=ci,value=h),bg="0D2040",fg=_BLUE)
    ws.row_dimensions[3].height=20
    if not tareas_hoy:
        ws.merge_cells("A4:G4"); e=ws["A4"]; e.value="No hay tareas programadas para hoy."
        e.fill=_xf(_DARK2); e.font=_xfn(color=_TXT_M,italic=True); e.alignment=_xal(); e.border=_xb()
    else:
        for ri,t in enumerate(tareas_hoy,start=4):
            done=t["completada"]==1; est="✔  Completada" if done else "●  Pendiente"; ec=_ACCENT if done else _AMBER; rbg="1A4A24" if done else "3D2E00"
            vals=[t["id"],t["codigo"] or "—",t["descripcion"],t["categoria"] or "General",est,t.get("username") or "—",""]
            for ci,val in enumerate(vals,start=1):
                c=ws.cell(row=ri,column=ci,value=val); c.fill=_xf(rbg if ci==5 else _DARK2)
                c.font=_xfn(bold=(ci==5),color=ec if ci==5 else(_TXT_M if ci in(1,6) else _TXT_W),size=9 if ci in(1,6) else 10)
                c.alignment=_xal(h="center" if ci in(1,5) else "left",wrap=(ci==3)); c.border=_xb()
            ws.row_dimensions[ri].height=16
    _xaw(ws); ws.column_dimensions["C"].width=45; ws.column_dimensions["G"].width=20; ws.freeze_panes="A4"; ws.auto_filter.ref=ws.dimensions

@app.route("/exportar")
@login_required
def exportar():
    hoy      = datetime.now().strftime("%Y-%m-%d")
    user_id  = session.get("user_id")
    es_admin = session.get("es_admin", 0) >= 1
    username = session.get("user", "sistema")
    conn     = get_connection()
    base_q   = "SELECT t.id,t.descripcion,t.categoria,t.fecha,t.completada,t.codigo,t.usuario_id,u.username FROM tareas t LEFT JOIN usuarios u ON t.usuario_id=u.id "
    if es_admin:
        todas      = conn.execute(base_q+"ORDER BY t.categoria,t.id").fetchall()
        tareas_hoy = conn.execute(base_q+"WHERE t.fecha=? ORDER BY t.id",(hoy,)).fetchall()
    else:
        todas      = conn.execute(base_q+"WHERE t.usuario_id=? ORDER BY t.categoria,t.id",(user_id,)).fetchall()
        tareas_hoy = conn.execute(base_q+"WHERE t.usuario_id=? AND t.fecha=? ORDER BY t.id",(user_id,hoy)).fetchall()
    conn.close()
    cats_dict = defaultdict(list)
    for t in todas: cats_dict[t["categoria"] or "General"].append(t)
    wb = Workbook()
    _hoja_resumen(wb, todas, hoy, es_admin, username)
    for idx,(cat_nombre,tareas_cat) in enumerate(sorted(cats_dict.items())):
        _hoja_categoria(wb, cat_nombre, tareas_cat, idx)
    _hoja_hoy(wb, tareas_hoy, hoy)
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Reporte_Tareas_{hoy}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── UTILIDADES ─────────────────────────────────────────────────────────────────

@app.route("/fix_passwords")
def fix_passwords():
    conn   = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, password FROM usuarios")
    users  = cursor.fetchall()
    migrados, ya_hash = [], []
    for u in users:
        pw = u["password"] or ""
        if pw and not pw.startswith("pbkdf2:") and not pw.startswith("scrypt:"):
            cursor.execute("UPDATE usuarios SET password=? WHERE id=?",
                           (generate_password_hash(pw), u["id"]))
            migrados.append(f"{u['username']} (id={u['id']})")
        else:
            ya_hash.append(u["username"])
    cursor.execute("UPDATE usuarios SET es_admin=2 WHERE username=?", ("admin",))
    conn.commit(); conn.close()
    lines = ["<h2>✅ Migración completada</h2>"]
    lines.append(f"<p><strong>Hasheados ahora:</strong> {len(migrados)}</p>")
    for m in migrados: lines.append(f"<p>🔑 {m}</p>")
    lines.append(f"<p><strong>Ya tenían hash:</strong> {len(ya_hash)} — {', '.join(ya_hash)}</p>")
    lines.append("<p>👑 admin → es_admin=2 (SuperAdmin)</p>")
    lines.append("<br><a href='/login'>👉 Ir al login</a>")
    return "".join(lines)


@app.route("/debug_login")
def debug_login():
    from db_mysql import get_db_name, MYSQL_CONFIG
    conn  = get_connection()
    db_real = conn.execute("SELECT DATABASE() AS db").fetchone()
    users = conn.execute("SELECT id, username, LEFT(password,30) as pw, es_admin FROM usuarios").fetchall()
    conn.close()
    lines = [f"<h2>BD activa: <b style='color:red'>{db_real['db']}</b> | .env MYSQL_DB={get_db_name()}</h2>"]
    lines.append("<table border=1><tr><th>id</th><th>user</th><th>pw_inicio</th><th>es_admin</th></tr>")
    for u in users:
        lines.append(f"<tr><td>{u['id']}</td><td>{u['username']}</td><td>{u['pw']}</td><td>{u['es_admin']}</td></tr>")
    lines.append("</table><br><a href='/fix_passwords'>👉 Hashear contraseñas</a>")
    return "".join(lines)


# ── ARRANCAR ───────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    inicializar_todo()
    inicializar_formacion()
    app.run(debug=True, port=5000)